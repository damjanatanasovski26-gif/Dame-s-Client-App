from flask import Flask, render_template, request, redirect, url_for, session, Response, send_file, abort, jsonify
from flask_migrate import Migrate
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func, or_
import click
from datetime import datetime, date, timedelta, timezone
import calendar
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.middleware.proxy_fix import ProxyFix
from functools import wraps
import base64
import csv
import hashlib
import hmac
import io
import json
import os
import secrets
import re
import shutil
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen
from math import ceil
from werkzeug.utils import secure_filename
import uuid

try:
    import pytesseract
except ImportError:
    pytesseract = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

from PIL import Image, ImageDraw, ImageFilter, ImageFont, ImageOps

try:
    import requests
except ImportError:
    requests = None

try:
    from bs4 import BeautifulSoup
except ImportError:
    BeautifulSoup = None


def load_local_env(env_path: str):
    if not os.path.exists(env_path):
        return
    try:
        with open(env_path, "r", encoding="utf-8") as fh:
            for raw_line in fh:
                line = raw_line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                key, value = line.split("=", 1)
                key = key.strip()
                value = value.strip().strip('"').strip("'")
                if key and key not in os.environ:
                    os.environ[key] = value
    except OSError:
        pass


load_local_env(os.path.join(os.path.dirname(__file__), ".env"))

app = Flask(__name__)
APP_ENV = os.environ.get("TRAINER_APP_ENV", os.environ.get("FLASK_ENV", "development")).lower()
IS_PROD = APP_ENV == "production"

secret_key = os.environ.get("TRAINER_APP_SECRET_KEY") or os.environ.get("SECRET_KEY")
if IS_PROD and not secret_key:
    raise RuntimeError("Missing TRAINER_APP_SECRET_KEY/SECRET_KEY in production.")
app.config["SECRET_KEY"] = secret_key or os.urandom(32)

database_uri = (
    os.environ.get("TRAINER_DATABASE_URI")
    or os.environ.get("DATABASE_URL")
    or "sqlite:///trainer.db"
)
# Railway Postgres URLs can use postgres://, but SQLAlchemy expects postgresql://.
if database_uri.startswith("postgres://"):
    database_uri = database_uri.replace("postgres://", "postgresql://", 1)
# Prefer psycopg v3 driver on PostgreSQL to avoid psycopg2/libpq runtime issues.
if database_uri.startswith("postgresql://"):
    database_uri = database_uri.replace("postgresql://", "postgresql+psycopg://", 1)
app.config["SQLALCHEMY_DATABASE_URI"] = database_uri
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["CSRF_ENABLED"] = True

# Session cookie hardening (production-safe, dev-friendly)
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = os.environ.get("SESSION_COOKIE_SAMESITE", "Lax")
app.config["SESSION_COOKIE_NAME"] = os.environ.get("SESSION_COOKIE_NAME", "trainer_session")
app.config["SESSION_COOKIE_SECURE"] = (
    os.environ.get("SESSION_COOKIE_SECURE", "1" if IS_PROD else "0").lower() in ("1", "true", "yes", "on")
)
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(
    days=int(os.environ.get("REMEMBER_ME_DAYS", "30"))
)
app.config["FORCE_HTTPS"] = (
    os.environ.get("FORCE_HTTPS", "1" if IS_PROD else "0").lower() in ("1", "true", "yes", "on")
)
app.config["ENABLE_SECURITY_HEADERS"] = True
app.config["LOGIN_MAX_ATTEMPTS"] = int(os.environ.get("LOGIN_MAX_ATTEMPTS", "5"))
app.config["LOGIN_WINDOW_SECONDS"] = int(os.environ.get("LOGIN_WINDOW_SECONDS", "300"))
app.config["LOGIN_LOCK_SECONDS"] = int(os.environ.get("LOGIN_LOCK_SECONDS", "600"))
app.config["UPLOAD_PROGRESS_DIR"] = os.environ.get(
    "UPLOAD_PROGRESS_DIR",
    os.path.join(app.root_path, "static", "uploads", "progress")
)
app.config["UPLOAD_LABEL_DIR"] = os.environ.get(
    "UPLOAD_LABEL_DIR",
    os.path.join(app.root_path, "static", "uploads", "nutrition_labels")
)
app.config["MAX_PHOTO_UPLOAD_BYTES"] = int(os.environ.get("MAX_PHOTO_UPLOAD_BYTES", str(10 * 1024 * 1024)))
app.config["USDA_API_KEY"] = os.environ.get("USDA_API_KEY", "").strip()
app.config["GOOGLE_VISION_API_KEY"] = os.environ.get("GOOGLE_VISION_API_KEY", "").strip()
app.config["GOOGLE_VISION_FEATURE_TYPE"] = os.environ.get("GOOGLE_VISION_FEATURE_TYPE", "DOCUMENT_TEXT_DETECTION").strip()
app.config["GOOGLE_VISION_LANGUAGE_HINTS"] = os.environ.get("GOOGLE_VISION_LANGUAGE_HINTS", "en,mk,sq").strip()
app.config["TESSERACT_CMD"] = os.environ.get("TESSERACT_CMD", "").strip()
app.config["TESSERACT_LANGS"] = os.environ.get("TESSERACT_LANGS", "eng+mkd+sqi+deu").strip()

os.makedirs(app.config["UPLOAD_PROGRESS_DIR"], exist_ok=True)
os.makedirs(app.config["UPLOAD_LABEL_DIR"], exist_ok=True)

if os.environ.get("TRUST_PROXY", "1").lower() in ("1", "true", "yes", "on"):
    app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)

db = SQLAlchemy(app)
migrate = Migrate(app, db)


@app.route("/manifest.webmanifest")
def pwa_manifest():
    return app.send_static_file("manifest.webmanifest")


@app.route("/service-worker.js")
def pwa_service_worker():
    resp = app.send_static_file("service-worker.js")
    resp.headers["Content-Type"] = "application/javascript; charset=utf-8"
    resp.headers["Cache-Control"] = "no-cache"
    return resp


def utc_now():
    # Use naive UTC timestamps because DB DateTime columns are timezone-naive.
    return datetime.utcnow()


def get_csrf_token():
    token = session.get("_csrf_token")
    if not token:
        token = secrets.token_urlsafe(32)
        session["_csrf_token"] = token
    return token


def truthy(v: str):
    return (v or "").lower() in ("1", "true", "yes", "on")


def log_security_event(action: str, details: str = ""):
    user = session.get("user_id")
    role = session.get("role")
    app.logger.info("[security] action=%s user_id=%s role=%s ip=%s details=%s", action, user, role, request.remote_addr, details)


def humanize_last_seen(ts: datetime | None):
    if not ts:
        return "-"
    dt = ts
    if dt.tzinfo is not None:
        dt = dt.astimezone(timezone.utc).replace(tzinfo=None)
    delta = utc_now() - dt
    seconds = int(max(delta.total_seconds(), 0))
    if seconds < 60:
        return "just now"
    if seconds < 3600:
        mins = seconds // 60
        return f"{mins} min ago"
    if seconds < 86400:
        hours = seconds // 3600
        return f"{hours}h ago"
    if seconds < 604800:
        days = seconds // 86400
        return f"{days}d ago"
    return dt.strftime("%d/%m/%Y %H:%M UTC")


def login_throttle_keys(username: str):
    ip = (request.remote_addr or "unknown").strip()
    normalized_username = (username or "").strip().lower() or "*"
    return [f"{ip}::{normalized_username}", f"{ip}::*"]


@app.context_processor
def inject_csrf_token():
    return {"csrf_token": get_csrf_token}


def asset_v(static_path: str) -> str:
    """Cache-busting version derived from the file's own mtime, so edits are
    always picked up by the browser without needing a manually-bumped
    query string."""
    full_path = os.path.join(app.static_folder, static_path)
    try:
        return str(int(os.path.getmtime(full_path)))
    except OSError:
        return "0"


@app.context_processor
def inject_asset_v():
    return {"asset_v": asset_v}


@app.template_filter("mkd")
def format_mkd(value):
    try:
        value = round(float(value))
    except (TypeError, ValueError):
        return value
    return f"{value:,}"


@app.before_request
def touch_last_seen():
    uid = session.get("user_id")
    if not uid:
        return
    if request.endpoint == "static":
        return
    user = db.session.get(User, uid)
    if not user:
        return
    now = utc_now()
    if user.last_seen_at and (now - user.last_seen_at).total_seconds() < 60:
        return
    user.last_seen_at = now
    db.session.commit()


@app.before_request
def csrf_protect():
    if not app.config.get("CSRF_ENABLED", True):
        return
    if request.method != "POST":
        return

    expected = session.get("_csrf_token")
    provided = request.form.get("csrf_token") or request.headers.get("X-CSRF-Token")
    if not expected or not provided or not hmac.compare_digest(expected, provided):
        abort(400, description="Invalid CSRF token.")


@app.before_request
def enforce_https():
    if not app.config.get("FORCE_HTTPS"):
        return
    if app.testing:
        return
    if request.path.startswith("/ping"):
        return
    is_https = request.is_secure or request.headers.get("X-Forwarded-Proto", "").lower() == "https"
    if not is_https:
        return redirect(request.url.replace("http://", "https://", 1), code=301)


@app.after_request
def apply_security_headers(resp):
    if not app.config.get("ENABLE_SECURITY_HEADERS", True):
        return resp
    resp.headers["X-Frame-Options"] = "DENY"
    resp.headers["X-Content-Type-Options"] = "nosniff"
    resp.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    resp.headers["Permissions-Policy"] = "geolocation=(), microphone=(), camera=()"
    if app.config.get("SESSION_COOKIE_SECURE"):
        resp.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
    csp = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline'; "
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
        "font-src 'self' https://fonts.gstatic.com; "
        "img-src 'self' data:; "
        "connect-src 'self'; "
        "frame-ancestors 'none'; "
        "base-uri 'self'; "
        "form-action 'self'"
    )
    resp.headers["Content-Security-Policy"] = csp
    return resp

# =========================
# Models
# =========================
class Client(db.Model):
    id = db.Column(db.Integer, primary_key=True)

    name = db.Column(db.String(100), nullable=False)
    phone = db.Column(db.String(50))
    plan = db.Column(db.String(50))

    # fallback if no payments exist
    weekly_sessions = db.Column(db.Integer, default=0)

    created_at = db.Column(db.DateTime, default=utc_now)
    is_active = db.Column(db.Boolean, nullable=False, default=True)
    daily_calorie_target = db.Column(db.Integer, nullable=True)

    # rollover system
    rollover_bonus = db.Column(db.Integer, default=0)  # bonus sessions for a specific next week
    rollover_for_week = db.Column(db.Date, nullable=True)  # week_start date for which rollover applies
    last_transfer_week = db.Column(db.Date, nullable=True)  # prevent multiple transfers per week


class Measurement(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    client_id = db.Column(db.Integer, db.ForeignKey("client.id"), nullable=False)

    date = db.Column(db.DateTime, default=utc_now)

    weight = db.Column(db.Float)
    chest = db.Column(db.Float)
    waist = db.Column(db.Float)
    stomach = db.Column(db.Float)
    glutes = db.Column(db.Float)
    arm_left = db.Column(db.Float)
    arm_right = db.Column(db.Float)
    quad_left = db.Column(db.Float)
    quad_right = db.Column(db.Float)
    calf_left = db.Column(db.Float)
    calf_right = db.Column(db.Float)


class SessionLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    client_id = db.Column(db.Integer, db.ForeignKey("client.id"), nullable=False)
    date = db.Column(db.DateTime, default=utc_now)
    note = db.Column(db.String(200))


class ClientNote(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    client_id = db.Column(db.Integer, db.ForeignKey("client.id"), nullable=False, index=True)
    created_at = db.Column(db.DateTime, default=utc_now, nullable=False)
    text = db.Column(db.String(500), nullable=False)
    is_private = db.Column(db.Boolean, nullable=False, default=False)
    created_by_role = db.Column(db.String(20), nullable=False, default="admin")


class Appointment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    client_id = db.Column(db.Integer, db.ForeignKey("client.id"), nullable=False, index=True)
    scheduled_for = db.Column(db.DateTime, nullable=False)
    status = db.Column(db.String(20), nullable=False, default="requested")  # requested/confirmed/completed/cancelled
    note = db.Column(db.String(200))
    created_at = db.Column(db.DateTime, default=utc_now, nullable=False)
    created_by_role = db.Column(db.String(20), nullable=False, default="client")


class ProgressPhoto(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    client_id = db.Column(db.Integer, db.ForeignKey("client.id"), nullable=False, index=True)
    created_at = db.Column(db.DateTime, default=utc_now, nullable=False)
    file_name = db.Column(db.String(255), nullable=False)
    note = db.Column(db.String(200))
    uploaded_by_role = db.Column(db.String(20), nullable=False, default="client")


class ClientGoal(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    client_id = db.Column(db.Integer, db.ForeignKey("client.id"), nullable=False, index=True)
    title = db.Column(db.String(120), nullable=False)
    goal_type = db.Column(db.String(20), nullable=False, default="custom")
    unit = db.Column(db.String(20), nullable=True)
    target_value = db.Column(db.Float, nullable=True)
    current_value = db.Column(db.Float, nullable=True)
    target_date = db.Column(db.Date, nullable=True)
    status = db.Column(db.String(20), nullable=False, default="active")  # active/completed/paused
    created_at = db.Column(db.DateTime, default=utc_now, nullable=False)
    note = db.Column(db.String(300))


class FoodItem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    client_id = db.Column(db.Integer, db.ForeignKey("client.id"), nullable=True, index=True)
    name = db.Column(db.String(120), nullable=False)
    brand = db.Column(db.String(120))
    serving_label = db.Column(db.String(80))
    source = db.Column(db.String(30), nullable=False, default="manual")
    source_ref = db.Column(db.String(120))
    barcode = db.Column(db.String(64))
    calories_per_100g = db.Column(db.Float, nullable=False)
    protein_per_100g = db.Column(db.Float, nullable=False, default=0)
    carbs_per_100g = db.Column(db.Float, nullable=False, default=0)
    fat_per_100g = db.Column(db.Float, nullable=False, default=0)
    created_at = db.Column(db.DateTime, default=utc_now, nullable=False)


class FoodLogEntry(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    client_id = db.Column(db.Integer, db.ForeignKey("client.id"), nullable=False, index=True)
    food_id = db.Column(db.Integer, db.ForeignKey("food_item.id"), nullable=False, index=True)
    logged_for = db.Column(db.Date, nullable=False, default=date.today, index=True)
    meal_type = db.Column(db.String(20), nullable=False, default="snack")
    quantity_grams = db.Column(db.Float, nullable=False)
    food_name = db.Column(db.String(160), nullable=False)
    calories = db.Column(db.Float, nullable=False, default=0)
    protein = db.Column(db.Float, nullable=False, default=0)
    carbs = db.Column(db.Float, nullable=False, default=0)
    fat = db.Column(db.Float, nullable=False, default=0)
    note = db.Column(db.String(200))
    created_at = db.Column(db.DateTime, default=utc_now, nullable=False)


class StrengthLogEntry(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    client_id = db.Column(db.Integer, db.ForeignKey("client.id"), nullable=False, index=True)
    source = db.Column(db.String(30), nullable=False, default="lyfta")
    workout_title = db.Column(db.String(120))
    logged_at = db.Column(db.DateTime, nullable=False, index=True)
    exercise = db.Column(db.String(160), nullable=False, index=True)
    weight = db.Column(db.Float, nullable=True)
    reps = db.Column(db.Integer, nullable=True)
    rir_rpe = db.Column(db.String(20))
    duration = db.Column(db.String(20))
    set_type = db.Column(db.String(30))
    source_file = db.Column(db.String(255))
    created_at = db.Column(db.DateTime, default=utc_now, nullable=False)


class Payment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    client_id = db.Column(db.Integer, db.ForeignKey("client.id"), nullable=False)

    paid_on = db.Column(db.DateTime, default=utc_now)

    # start date of coverage (DD/MM/YYYY from form)
    start_date = db.Column(db.Date, nullable=False)

    months = db.Column(db.Integer, nullable=False, default=1)
    sessions_per_week = db.Column(db.Integer, nullable=False, default=3)

    # monthly price used (5000 or 7000)
    monthly_price = db.Column(db.Integer, nullable=False, default=5000)

    # total paid (e.g. 10000)
    amount_paid = db.Column(db.Integer, nullable=False, default=0)
    due_date_override = db.Column(db.Date, nullable=True)

    note = db.Column(db.String(200))


class BudgetWeek(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    week_start = db.Column(db.Date, unique=True, nullable=False, index=True)
    weekly_budget = db.Column(db.Integer, nullable=False, default=0)


class BudgetDay(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    day = db.Column(db.Date, unique=True, nullable=False, index=True)
    spent = db.Column(db.Integer, nullable=False, default=0)


class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)

    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)

    # "admin" or "client"
    role = db.Column(db.String(20), default="client", nullable=False)

    # client users link to a Client profile
    client_id = db.Column(db.Integer, db.ForeignKey("client.id"), nullable=True)
    must_change_password = db.Column(db.Boolean, nullable=False, default=False)
    last_login_at = db.Column(db.DateTime, nullable=True)
    last_seen_at = db.Column(db.DateTime, nullable=True)


class LoginThrottle(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    key = db.Column(db.String(255), unique=True, nullable=False, index=True)
    count = db.Column(db.Integer, nullable=False, default=0)
    first_ts = db.Column(db.DateTime, nullable=False, default=utc_now)
    lock_until = db.Column(db.DateTime, nullable=True)
    updated_at = db.Column(db.DateTime, nullable=False, default=utc_now, onupdate=utc_now)


# =========================
# Helpers
# =========================
def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login"))
        return view(*args, **kwargs)
    return wrapped


def is_admin():
    return session.get("role") == "admin"


def current_client_id():
    return session.get("client_id")


def current_user():
    uid = session.get("user_id")
    if not uid:
        return None
    return User.query.filter_by(id=uid).first()


def parse_ddmmyyyy(s: str):
    s = (s or "").strip()
    # expects DD/MM/YYYY
    try:
        dd, mm, yy = s.split("/")
        return date(int(yy), int(mm), int(dd))
    except Exception:
        return None


def to_float(value):
    value = (value or "").strip().replace(",", ".")
    if value == "":
        return None
    try:
        return float(value)
    except ValueError:
        return None


MEASUREMENT_FIELDS = [
    "weight",
    "chest",
    "waist",
    "stomach",
    "glutes",
    "arm_left",
    "arm_right",
    "quad_left",
    "quad_right",
    "calf_left",
    "calf_right",
]

MEAL_TYPES = ("breakfast", "lunch", "dinner", "snack")
LABEL_REGEX_PATTERNS = {
    "calories": [
        r"(?:energy|energia|energjia|vlera\s+energjetike|energetska\s+vrednost|\u0435\u043d\u0435\u0440\u0433\u0438\u0458\u0430|\u0435\u043d\u0435\u0440\u0433\u0435\u0442\u0441\u043a\u0430\s+\u0432\u0440\u0435\u0434\u043d\u043e\u0441\u0442)[^\n\r]{0,50}?(\d+(?:[.,]\d+)?)\s*kcal",
        r"(\d+(?:[.,]\d+)?)\s*kcal",
    ],
    "protein": [
        r"(?:protein(?:s)?|proteina|proteini|\u0431\u0435\u043b\u043a\u043e\u0432\u0438\u043d\u0438|\u043f\u0440\u043e\u0442\u0435\u0438\u043d\u0438)[^\n\r]{0,40}?(\d+(?:[.,]\d+)?)\s*g",
    ],
    "carbs": [
        r"(?:carbohydrate(?:s)?|carbs|karbohidrate(?:t)?|ugljeni\s+hidrati|\u0458\u0430\u0433\u043b\u0435\u0445\u0438\u0434\u0440\u0430\u0442\u0438)[^\n\r]{0,40}?(\d+(?:[.,]\d+)?)\s*g",
    ],
    "fat": [
        r"(?:fat|fats|yndyr(?:e|a|na)|yndyr\u00eb|yndyrna|mast(?:i)?|\u043c\u0430\u0441\u0442\u0438)[^\n\r]{0,40}?(\d+(?:[.,]\d+)?)\s*g",
    ],
}
RAW_INGREDIENT_LIBRARY = [
    {"slug": "apple", "name": "Apple", "group": "Shared Raw Ingredient", "calories": 52, "protein": 0.3, "carbs": 13.8, "fat": 0.2},
    {"slug": "banana", "name": "Banana", "group": "Shared Raw Ingredient", "calories": 89, "protein": 1.1, "carbs": 22.8, "fat": 0.3},
    {"slug": "orange", "name": "Orange", "group": "Shared Raw Ingredient", "calories": 47, "protein": 0.9, "carbs": 11.8, "fat": 0.1},
    {"slug": "pear", "name": "Pear", "group": "Shared Raw Ingredient", "calories": 57, "protein": 0.4, "carbs": 15.2, "fat": 0.1},
    {"slug": "peach", "name": "Peach", "group": "Shared Raw Ingredient", "calories": 39, "protein": 0.9, "carbs": 9.5, "fat": 0.3},
    {"slug": "kiwi", "name": "Kiwi", "group": "Shared Raw Ingredient", "calories": 61, "protein": 1.1, "carbs": 14.7, "fat": 0.5},
    {"slug": "pineapple", "name": "Pineapple", "group": "Shared Raw Ingredient", "calories": 50, "protein": 0.5, "carbs": 13.1, "fat": 0.1},
    {"slug": "grapes", "name": "Grapes", "group": "Shared Raw Ingredient", "calories": 69, "protein": 0.7, "carbs": 18.1, "fat": 0.2},
    {"slug": "strawberries", "name": "Strawberries", "group": "Shared Raw Ingredient", "calories": 32, "protein": 0.7, "carbs": 7.7, "fat": 0.3},
    {"slug": "blueberries", "name": "Blueberries", "group": "Shared Raw Ingredient", "calories": 57, "protein": 0.7, "carbs": 14.5, "fat": 0.3},
    {"slug": "watermelon", "name": "Watermelon", "group": "Shared Raw Ingredient", "calories": 30, "protein": 0.6, "carbs": 7.6, "fat": 0.2},
    {"slug": "lemon", "name": "Lemon", "group": "Shared Raw Ingredient", "calories": 29, "protein": 1.1, "carbs": 9.3, "fat": 0.3},
    {"slug": "avocado", "name": "Avocado", "group": "Shared Raw Ingredient", "calories": 160, "protein": 2.0, "carbs": 8.5, "fat": 14.7},
    {"slug": "tomato", "name": "Tomato", "group": "Shared Raw Ingredient", "calories": 18, "protein": 0.9, "carbs": 3.9, "fat": 0.2},
    {"slug": "cucumber", "name": "Cucumber", "group": "Shared Raw Ingredient", "calories": 15, "protein": 0.7, "carbs": 3.6, "fat": 0.1},
    {"slug": "carrot", "name": "Carrot", "group": "Shared Raw Ingredient", "calories": 41, "protein": 0.9, "carbs": 9.6, "fat": 0.2},
    {"slug": "broccoli", "name": "Broccoli", "group": "Shared Raw Ingredient", "calories": 34, "protein": 2.8, "carbs": 6.6, "fat": 0.4},
    {"slug": "cauliflower", "name": "Cauliflower", "group": "Shared Raw Ingredient", "calories": 25, "protein": 1.9, "carbs": 5.0, "fat": 0.3},
    {"slug": "spinach", "name": "Spinach", "group": "Shared Raw Ingredient", "calories": 23, "protein": 2.9, "carbs": 3.6, "fat": 0.4},
    {"slug": "lettuce", "name": "Lettuce", "group": "Shared Raw Ingredient", "calories": 15, "protein": 1.4, "carbs": 2.9, "fat": 0.2},
    {"slug": "onion", "name": "Onion", "group": "Shared Raw Ingredient", "calories": 40, "protein": 1.1, "carbs": 9.3, "fat": 0.1},
    {"slug": "garlic", "name": "Garlic", "group": "Shared Raw Ingredient", "calories": 149, "protein": 6.4, "carbs": 33.1, "fat": 0.5},
    {"slug": "bell-pepper", "name": "Bell Pepper", "group": "Shared Raw Ingredient", "calories": 31, "protein": 1.0, "carbs": 6.0, "fat": 0.3},
    {"slug": "zucchini", "name": "Zucchini", "group": "Shared Raw Ingredient", "calories": 17, "protein": 1.2, "carbs": 3.1, "fat": 0.3},
    {"slug": "mushrooms", "name": "Mushrooms", "group": "Shared Raw Ingredient", "calories": 22, "protein": 3.1, "carbs": 3.3, "fat": 0.3},
    {"slug": "cabbage", "name": "Cabbage", "group": "Shared Raw Ingredient", "calories": 25, "protein": 1.3, "carbs": 5.8, "fat": 0.1},
    {"slug": "green-beans", "name": "Green Beans", "group": "Shared Raw Ingredient", "calories": 31, "protein": 1.8, "carbs": 7.0, "fat": 0.2},
    {"slug": "peas", "name": "Peas", "group": "Shared Raw Ingredient", "calories": 81, "protein": 5.4, "carbs": 14.5, "fat": 0.4},
    {"slug": "corn", "name": "Corn", "group": "Shared Raw Ingredient", "calories": 86, "protein": 3.3, "carbs": 18.7, "fat": 1.4},
    {"slug": "potato", "name": "Potato", "group": "Shared Raw Ingredient", "calories": 77, "protein": 2.0, "carbs": 17.5, "fat": 0.1},
    {"slug": "sweet-potato", "name": "Sweet Potato", "group": "Shared Raw Ingredient", "calories": 86, "protein": 1.6, "carbs": 20.1, "fat": 0.1},
    {"slug": "rice-white-dry", "name": "Rice, White, Dry", "group": "Shared Raw Ingredient", "calories": 365, "protein": 7.1, "carbs": 80.0, "fat": 0.7},
    {"slug": "rice-brown-dry", "name": "Rice, Brown, Dry", "group": "Shared Raw Ingredient", "calories": 370, "protein": 7.9, "carbs": 77.2, "fat": 2.9},
    {"slug": "oats-dry", "name": "Oats, Dry", "group": "Shared Raw Ingredient", "calories": 389, "protein": 16.9, "carbs": 66.3, "fat": 6.9},
    {"slug": "quinoa-dry", "name": "Quinoa, Dry", "group": "Shared Raw Ingredient", "calories": 368, "protein": 14.1, "carbs": 64.2, "fat": 6.1},
    {"slug": "barley-dry", "name": "Barley, Dry", "group": "Shared Raw Ingredient", "calories": 354, "protein": 12.5, "carbs": 73.5, "fat": 2.3},
    {"slug": "whole-wheat-pasta-dry", "name": "Whole Wheat Pasta, Dry", "group": "Shared Raw Ingredient", "calories": 348, "protein": 14.6, "carbs": 72.0, "fat": 2.5},
    {"slug": "lentils-dry", "name": "Lentils, Dry", "group": "Shared Raw Ingredient", "calories": 353, "protein": 25.8, "carbs": 60.1, "fat": 1.1},
    {"slug": "chickpeas-dry", "name": "Chickpeas, Dry", "group": "Shared Raw Ingredient", "calories": 364, "protein": 19.3, "carbs": 60.7, "fat": 6.0},
    {"slug": "black-beans-dry", "name": "Black Beans, Dry", "group": "Shared Raw Ingredient", "calories": 341, "protein": 21.6, "carbs": 62.4, "fat": 1.4},
    {"slug": "kidney-beans-dry", "name": "Kidney Beans, Dry", "group": "Shared Raw Ingredient", "calories": 333, "protein": 23.6, "carbs": 60.0, "fat": 0.8},
    {"slug": "white-beans-dry", "name": "White Beans, Dry", "group": "Shared Raw Ingredient", "calories": 336, "protein": 21.1, "carbs": 61.3, "fat": 1.6},
    {"slug": "olive-oil", "name": "Olive Oil", "group": "Shared Raw Ingredient", "calories": 884, "protein": 0.0, "carbs": 0.0, "fat": 100.0},
    {"slug": "sunflower-oil", "name": "Sunflower Oil", "group": "Shared Raw Ingredient", "calories": 884, "protein": 0.0, "carbs": 0.0, "fat": 100.0},
    {"slug": "canola-oil", "name": "Canola Oil", "group": "Shared Raw Ingredient", "calories": 884, "protein": 0.0, "carbs": 0.0, "fat": 100.0},
    {"slug": "coconut-oil", "name": "Coconut Oil", "group": "Shared Raw Ingredient", "calories": 892, "protein": 0.0, "carbs": 0.0, "fat": 100.0},
    {"slug": "almonds", "name": "Almonds", "group": "Shared Raw Ingredient", "calories": 579, "protein": 21.2, "carbs": 21.6, "fat": 49.9},
    {"slug": "walnuts", "name": "Walnuts", "group": "Shared Raw Ingredient", "calories": 654, "protein": 15.2, "carbs": 13.7, "fat": 65.2},
    {"slug": "cashews", "name": "Cashews", "group": "Shared Raw Ingredient", "calories": 553, "protein": 18.2, "carbs": 30.2, "fat": 43.8},
    {"slug": "peanuts", "name": "Peanuts", "group": "Shared Raw Ingredient", "calories": 567, "protein": 25.8, "carbs": 16.1, "fat": 49.2},
    {"slug": "pistachios", "name": "Pistachios", "group": "Shared Raw Ingredient", "calories": 562, "protein": 20.2, "carbs": 28.0, "fat": 45.3},
    {"slug": "chia-seeds", "name": "Chia Seeds", "group": "Shared Raw Ingredient", "calories": 486, "protein": 16.5, "carbs": 42.1, "fat": 30.7},
    {"slug": "flax-seeds", "name": "Flax Seeds", "group": "Shared Raw Ingredient", "calories": 534, "protein": 18.3, "carbs": 28.9, "fat": 42.2},
    {"slug": "pumpkin-seeds", "name": "Pumpkin Seeds", "group": "Shared Raw Ingredient", "calories": 559, "protein": 30.2, "carbs": 10.7, "fat": 49.0},
    {"slug": "sunflower-seeds", "name": "Sunflower Seeds", "group": "Shared Raw Ingredient", "calories": 584, "protein": 20.8, "carbs": 20.0, "fat": 51.5},
    {"slug": "sesame-seeds", "name": "Sesame Seeds", "group": "Shared Raw Ingredient", "calories": 573, "protein": 17.7, "carbs": 23.4, "fat": 49.7},
    {"slug": "chicken-breast-raw", "name": "Chicken Breast, Raw", "group": "Shared Raw Ingredient", "calories": 120, "protein": 22.5, "carbs": 0.0, "fat": 2.6},
    {"slug": "chicken-thigh-raw", "name": "Chicken Thigh, Raw", "group": "Shared Raw Ingredient", "calories": 144, "protein": 16.8, "carbs": 0.0, "fat": 8.1},
    {"slug": "ground-beef-lean-raw", "name": "Ground Beef, Lean, Raw", "group": "Shared Raw Ingredient", "calories": 176, "protein": 20.0, "carbs": 0.0, "fat": 10.0},
    {"slug": "ground-turkey-raw", "name": "Ground Turkey, Raw", "group": "Shared Raw Ingredient", "calories": 149, "protein": 19.0, "carbs": 0.0, "fat": 8.0},
    {"slug": "salmon-raw", "name": "Salmon, Raw", "group": "Shared Raw Ingredient", "calories": 208, "protein": 20.4, "carbs": 0.0, "fat": 13.4},
    {"slug": "tuna-raw", "name": "Tuna, Raw", "group": "Shared Raw Ingredient", "calories": 132, "protein": 28.0, "carbs": 0.0, "fat": 1.3},
    {"slug": "pork-loin-raw", "name": "Pork Loin, Raw", "group": "Shared Raw Ingredient", "calories": 143, "protein": 21.0, "carbs": 0.0, "fat": 6.0},
    {"slug": "egg-whole-raw", "name": "Egg, Whole, Raw", "group": "Shared Raw Ingredient", "calories": 143, "protein": 12.6, "carbs": 0.7, "fat": 9.5},
    {"slug": "egg-white-raw", "name": "Egg White, Raw", "group": "Shared Raw Ingredient", "calories": 52, "protein": 10.9, "carbs": 0.7, "fat": 0.2},
]


def parse_measurement_form(form_data):
    parsed = {}
    invalid_fields = []
    for field in MEASUREMENT_FIELDS:
        raw = (form_data.get(field) or "").strip().replace(",", ".")
        if raw == "":
            parsed[field] = None
            continue
        try:
            parsed[field] = float(raw)
        except ValueError:
            invalid_fields.append(field.replace("_", " "))

    if invalid_fields:
        return None, f"Invalid number for: {', '.join(invalid_fields)}."
    if all(v is None for v in parsed.values()):
        return None, "Enter at least one measurement value."
    return parsed, None


def normalize_goal_type(value: str | None):
    v = (value or "").strip().lower()
    if v not in ("weight", "measurement", "habit", "performance", "custom"):
        return "custom"
    return v


def normalize_goal_unit(value: str | None):
    unit = (value or "").strip()
    if not unit:
        return None
    return unit[:20]


def default_goal_unit(goal_type: str):
    if goal_type == "weight":
        return "kg"
    return None


def is_weight_goal(goal: ClientGoal) -> bool:
    if normalize_goal_type(getattr(goal, "goal_type", None)) == "weight":
        return True
    text = f"{goal.title or ''} {goal.note or ''}".lower()
    markers = ("kg", "kilo", "weight", "tezina", "тежина")
    return any(marker in text for marker in markers)


def get_latest_weight_value(client_id: int):
    latest_weight = (
        Measurement.query.filter_by(client_id=client_id)
        .filter(Measurement.weight.isnot(None))
        .order_by(Measurement.date.desc(), Measurement.id.desc())
        .first()
    )
    return latest_weight.weight if latest_weight else None


def sync_weight_goal_progress(client_id: int):
    latest_weight = get_latest_weight_value(client_id)
    if latest_weight is None:
        return 0

    goals = (
        ClientGoal.query.filter_by(client_id=client_id)
        .filter(ClientGoal.status != "completed")
        .all()
    )
    updated = 0
    for goal in goals:
        if not is_weight_goal(goal):
            continue
        if goal.current_value != latest_weight:
            goal.current_value = latest_weight
            updated += 1
    return updated


def to_int(value, default=0):
    value = (value or "").strip()
    if value == "":
        return default
    try:
        return int(value)
    except ValueError:
        return default


def parse_datetime_local(value: str):
    value = (value or "").strip()
    if not value:
        return None
    for fmt in ("%Y-%m-%dT%H:%M", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(value, fmt)
        except Exception:
            continue
    return None


def parse_iso_date(value: str):
    value = (value or "").strip()
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except Exception:
        return None


def parse_iso_datetime(value: str):
    value = (value or "").strip()
    if not value:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(value, fmt)
        except Exception:
            continue
    return None


def parse_strength_export_datetime(value):
    if isinstance(value, datetime):
        return value
    value = str(value or "").strip()
    if not value:
        return None
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%d %b %Y, %H:%M",
        "%d %B %Y, %H:%M",
    ):
        try:
            return datetime.strptime(value, fmt)
        except Exception:
            continue
    return None


def normalize_meal_type(value: str | None):
    meal_type = (value or "").strip().lower()
    if meal_type not in MEAL_TYPES:
        return "snack"
    return meal_type


def meal_type_label(meal_type: str):
    return normalize_meal_type(meal_type).title()


def estimate_one_rep_max(weight: float | None, reps: int | None):
    if weight is None or reps is None or reps <= 0:
        return None
    if reps == 1:
        return round(weight, 1)
    return round(weight * (1 + (reps / 30)), 1)


def resolve_tesseract_cmd():
    configured = app.config.get("TESSERACT_CMD", "").strip()
    if configured:
        return configured
    return shutil.which("tesseract") or ""


def google_vision_enabled():
    return bool(app.config.get("GOOGLE_VISION_API_KEY", "").strip())


def tesseract_scan_enabled():
    if pytesseract is None:
        return False
    tesseract_cmd = resolve_tesseract_cmd()
    if not tesseract_cmd:
        return False
    try:
        pytesseract.pytesseract.tesseract_cmd = tesseract_cmd
        pytesseract.get_tesseract_version()
        return True
    except Exception:
        return False


def label_scan_enabled():
    return google_vision_enabled() or tesseract_scan_enabled()


def label_scan_provider_label():
    if google_vision_enabled():
        return "Google Vision OCR"
    if tesseract_scan_enabled():
        return "Tesseract OCR"
    return ""


def get_tesseract_languages():
    requested = [lang.strip() for lang in app.config.get("TESSERACT_LANGS", "").split("+") if lang.strip()]
    if not requested:
        return None
    try:
        installed = set(pytesseract.get_languages(config=""))
    except Exception:
        installed = set()
    available = [lang for lang in requested if lang in installed]
    if available:
        return "+".join(available)
    return "+".join(requested)


def parse_positive_float(value, label: str, *, allow_zero: bool = False):
    parsed = to_float(value)
    if parsed is None:
        return None, f"{label} is required."
    if parsed < 0 or (not allow_zero and parsed == 0):
        comparator = "0 or more" if allow_zero else "more than 0"
        return None, f"{label} must be {comparator}."
    return parsed, None


def nutrition_redirect(client_id: int, *, msg: str | None = None, err: str | None = None, logged_for: date | None = None, anchor: str | None = None):
    params = {"client_id": client_id, "tab": "nutrition"}
    if msg:
        params["msg"] = msg
    if err:
        params["err"] = err
    if logged_for:
        params["nutrition_date"] = logged_for.isoformat()
    if anchor:
        params["_anchor"] = anchor
    return redirect(url_for("client_profile", **params))


def nutrition_search_session_key(client_id: int):
    return f"nutrition_search_results_{client_id}"


def nutrition_label_draft_session_key(client_id: int):
    return f"nutrition_label_draft_{client_id}"


def set_nutrition_search_results(client_id: int, results: list[dict]):
    session[nutrition_search_session_key(client_id)] = results


def get_nutrition_search_results(client_id: int):
    return session.get(nutrition_search_session_key(client_id), [])


def set_nutrition_label_draft(client_id: int, draft: dict | None):
    key = nutrition_label_draft_session_key(client_id)
    if draft:
        session[key] = draft
    else:
        session.pop(key, None)
    session.modified = True


def get_nutrition_label_draft(client_id: int):
    return session.get(nutrition_label_draft_session_key(client_id), {})


def get_accessible_food(client_id: int, food_id: int):
    return (
        FoodItem.query
        .filter(FoodItem.id == food_id)
        .filter(or_(FoodItem.client_id.is_(None), FoodItem.client_id == client_id))
        .first()
    )


def parse_decimal_text(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = str(value).strip().replace(",", ".")
    if not cleaned:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def normalize_food_search_text(value: str):
    return re.sub(r"[^a-z0-9]+", " ", (value or "").strip().lower()).strip()


CAPNUTRA_BASE_URL = "http://104.155.19.23/capnutra"
CAPNUTRA_SOURCE_NAME = "CAPNUTRA Serbian FCDB"
CAPNUTRA_IMPORT_COMPONENTS = {
    "calories_per_100g": "",
    "protein_per_100g": "PROT",
    "carbs_per_100g": "CHO",
    "fat_per_100g": "FAT",
}


def clean_capnutra_text(value: str):
    text = (value or "").replace("\xa0", " ")
    return re.sub(r"\s+", " ", text).strip()


def parse_capnutra_food_rows(html: str):
    if BeautifulSoup is None:
        raise RuntimeError("beautifulsoup4 is required for CAPNUTRA import.")

    soup = BeautifulSoup(html or "", "html.parser")
    rows = []
    for tr in soup.find_all("tr", id=re.compile(r"^basic\d+$")):
        trigger = tr.find(attrs={"onclick": re.compile(r"detail\(")})
        if not trigger:
            trigger = tr.find("a", href=re.compile(r"detail\("))
        call_text = (trigger.get("onclick") if trigger and trigger.has_attr("onclick") else "") or (
            trigger.get("href") if trigger and trigger.has_attr("href") else ""
        )
        id_match = re.search(r"detail\('\d+'\s*,\s*'(\d+)'\)", call_text)
        name_node = tr.find("strong")
        if not id_match or not name_node:
            continue

        text = clean_capnutra_text(tr.get_text(" "))
        value_match = re.search(r"\b([0-9]+(?:[.,][0-9]+)?)\s*kcal\b", text, re.I)
        if not value_match:
            numeric_cells = [
                clean_capnutra_text(node.get_text(" "))
                for node in tr.find_all(["td", "div"])
                if parse_decimal_text(clean_capnutra_text(node.get_text(" "))) is not None
            ]
            value_text = numeric_cells[-1] if numeric_cells else ""
            value_match = re.search(r"^([0-9]+(?:[.,][0-9]+)?)$", value_text)

        value = parse_decimal_text(value_match.group(1)) if value_match else None
        rows.append({
            "source_ref": id_match.group(1),
            "name": clean_capnutra_text(name_node.get_text(" ")),
            "value": value,
        })
    return rows


def parse_capnutra_total_pages(payload: str):
    parts = payload.split("||")
    if len(parts) < 3:
        return 1
    try:
        return max(int(parts[2].strip()), 1)
    except ValueError:
        return 1


def fetch_capnutra_food_page(component: str = "", page: int = 0, itempage: int = 60):
    params = urlencode({
        "pageNum": page,
        "food_gr": "",
        "food_name": "",
        "itempage": itempage,
        "xsort": "",
        "xcomp": component,
    })
    url = f"{CAPNUTRA_BASE_URL}/food.php?{params}"
    if requests is not None:
        response = requests.get(url, timeout=20, headers={"User-Agent": "TrainerApp/1.0 (+food-library-import)"})
        response.raise_for_status()
        return response.text

    req = Request(
        url,
        headers={"User-Agent": "TrainerApp/1.0 (+food-library-import)"},
    )
    with urlopen(req, timeout=20) as resp:
        return resp.read().decode("utf-8", errors="replace")


def fetch_capnutra_food_library(progress=None):
    foods_by_ref = {}
    total_pages = None

    for field, component in CAPNUTRA_IMPORT_COMPONENTS.items():
        first_payload = fetch_capnutra_food_page(component=component, page=0)
        pages = parse_capnutra_total_pages(first_payload)
        if total_pages is None:
            total_pages = pages

        for row in parse_capnutra_food_rows(first_payload):
            foods_by_ref.setdefault(row["source_ref"], {
                "source_ref": row["source_ref"],
                "name": row["name"],
                "brand": CAPNUTRA_SOURCE_NAME,
                "calories_per_100g": 0.0,
                "protein_per_100g": 0.0,
                "carbs_per_100g": 0.0,
                "fat_per_100g": 0.0,
            })[field] = row["value"] or 0.0

        for page in range(2, pages + 1):
            payload = fetch_capnutra_food_page(component=component, page=page)
            for row in parse_capnutra_food_rows(payload):
                item = foods_by_ref.setdefault(row["source_ref"], {
                    "source_ref": row["source_ref"],
                    "name": row["name"],
                    "brand": CAPNUTRA_SOURCE_NAME,
                    "calories_per_100g": 0.0,
                    "protein_per_100g": 0.0,
                    "carbs_per_100g": 0.0,
                    "fat_per_100g": 0.0,
                })
                item["name"] = item["name"] or row["name"]
                item[field] = row["value"] or 0.0

            if progress:
                progress(field, page, pages, len(foods_by_ref))

    return [
        item for item in foods_by_ref.values()
        if item["name"] and item["calories_per_100g"] > 0
    ]


def import_capnutra_foods(food_rows: list[dict] | None = None, prune_missing: bool = False):
    rows = food_rows if food_rows is not None else fetch_capnutra_food_library()
    existing = {
        item.source_ref: item
        for item in FoodItem.query.filter_by(source="capnutra").all()
    }

    created = 0
    updated = 0
    seen_refs = set()
    for row in rows:
        ref = str(row.get("source_ref") or "").strip()
        name = (row.get("name") or "").strip()
        calories = float(row.get("calories_per_100g") or 0)
        if not ref or not name or calories <= 0:
            continue

        seen_refs.add(ref)
        food = existing.get(ref)
        if not food:
            food = FoodItem(client_id=None, source="capnutra", source_ref=ref)
            db.session.add(food)
            created += 1
        else:
            updated += 1

        food.client_id = None
        food.name = name[:120]
        food.brand = (row.get("brand") or CAPNUTRA_SOURCE_NAME)[:120]
        food.serving_label = "100g"
        food.barcode = None
        food.calories_per_100g = round(calories, 1)
        food.protein_per_100g = round(float(row.get("protein_per_100g") or 0), 1)
        food.carbs_per_100g = round(float(row.get("carbs_per_100g") or 0), 1)
        food.fat_per_100g = round(float(row.get("fat_per_100g") or 0), 1)

    removed = 0
    if food_rows is None or prune_missing:
        for ref, food in existing.items():
            if ref not in seen_refs:
                db.session.delete(food)
                removed += 1

    db.session.commit()
    return {"created": created, "updated": updated, "removed": removed, "total": len(seen_refs)}


def extract_nutrient_value(nutrients, names: tuple[str, ...]):
    for nutrient in nutrients or []:
        name = (nutrient.get("nutrientName") or nutrient.get("name") or "").strip().lower()
        unit = (nutrient.get("unitName") or nutrient.get("unit") or "").strip().lower()
        value = nutrient.get("value")
        if value is None:
            continue
        if name in names:
            if "energy" in name and unit not in ("kcal", "kcal/100g"):
                continue
            return float(value)
    return 0.0


def http_get_json(url: str, headers: dict | None = None):
    req = Request(url, headers=headers or {"User-Agent": "TrainerApp/1.0"})
    with urlopen(req, timeout=12) as resp:
        body = resp.read().decode("utf-8")
    return json.loads(body)


def fetch_usda_foods(query: str):
    api_key = app.config.get("USDA_API_KEY", "")
    if not api_key:
        raise RuntimeError("USDA_API_KEY is not configured.")
    params = urlencode({
        "api_key": api_key,
        "query": query,
        "pageSize": 10,
        "dataType": ["Foundation", "SR Legacy", "Branded"],
    }, doseq=True)
    payload = http_get_json(f"https://api.nal.usda.gov/fdc/v1/foods/search?{params}")
    results = []
    for food in payload.get("foods", []):
        nutrients = food.get("foodNutrients", [])
        calories = extract_nutrient_value(nutrients, ("energy",))
        protein = extract_nutrient_value(nutrients, ("protein",))
        carbs = extract_nutrient_value(nutrients, ("carbohydrate, by difference", "carbohydrate"))
        fat = extract_nutrient_value(nutrients, ("total lipid (fat)", "fat"))
        if calories <= 0:
            continue
        results.append({
            "name": (food.get("description") or "USDA Food").title(),
            "brand": food.get("brandOwner") or food.get("brandName") or "USDA",
            "source": "usda",
            "source_ref": str(food.get("fdcId") or ""),
            "barcode": "",
            "calories_per_100g": round(calories, 1),
            "protein_per_100g": round(protein, 1),
            "carbs_per_100g": round(carbs, 1),
            "fat_per_100g": round(fat, 1),
        })
    return results


def map_off_product(product: dict):
    nutriments = product.get("nutriments") or {}
    calories = parse_decimal_text(str(nutriments.get("energy-kcal_100g") or nutriments.get("energy-kcal") or ""))
    if calories is None:
        kj = parse_decimal_text(str(nutriments.get("energy-kj_100g") or ""))
        calories = round(kj / 4.184, 1) if kj else None
    protein = parse_decimal_text(str(nutriments.get("proteins_100g") or nutriments.get("proteins") or "")) or 0.0
    carbs = parse_decimal_text(str(nutriments.get("carbohydrates_100g") or nutriments.get("carbohydrates") or "")) or 0.0
    fat = parse_decimal_text(str(nutriments.get("fat_100g") or nutriments.get("fat") or "")) or 0.0
    if calories is None:
        return None
    return {
        "name": (product.get("product_name") or product.get("generic_name") or "Open Food Facts Product").strip(),
        "brand": (product.get("brands") or "Open Food Facts").strip(),
        "source": "openfoodfacts",
        "source_ref": (product.get("_id") or "").strip(),
        "barcode": (product.get("code") or "").strip(),
        "calories_per_100g": round(calories, 1),
        "protein_per_100g": round(protein, 1),
        "carbs_per_100g": round(carbs, 1),
        "fat_per_100g": round(fat, 1),
    }


def fetch_open_food_facts_foods(query: str):
    barcode_query = (query or "").strip()
    if barcode_query.isdigit():
        payload = http_get_json(f"https://world.openfoodfacts.org/api/v2/product/{barcode_query}.json")
        product = payload.get("product") or {}
        mapped = map_off_product(product)
        return [mapped] if mapped else []

    params = urlencode({
        "search_terms": query,
        "search_simple": 1,
        "action": "process",
        "json": 1,
        "page_size": 10,
    })
    payload = http_get_json(f"https://world.openfoodfacts.org/cgi/search.pl?{params}")
    results = []
    for product in payload.get("products", []):
        mapped = map_off_product(product)
        if mapped:
            results.append(mapped)
    return results


def seed_reference_foods():
    desired = {item["slug"]: item for item in RAW_INGREDIENT_LIBRARY}
    existing = {
        item.source_ref: item
        for item in FoodItem.query.filter_by(source="seed").all()
    }

    changed = 0
    for slug, food in existing.items():
        if slug in desired:
            continue
        db.session.delete(food)
        changed += 1

    for slug, item in desired.items():
        existing_item = existing.get(slug)
        if not existing_item:
            db.session.add(FoodItem(
                client_id=None,
                name=item["name"],
                brand=item["group"],
                source="seed",
                source_ref=slug,
                calories_per_100g=item["calories"],
                protein_per_100g=item["protein"],
                carbs_per_100g=item["carbs"],
                fat_per_100g=item["fat"],
            ))
            changed += 1
            continue

        existing_item.client_id = None
        existing_item.name = item["name"]
        existing_item.brand = item["group"]
        existing_item.calories_per_100g = item["calories"]
        existing_item.protein_per_100g = item["protein"]
        existing_item.carbs_per_100g = item["carbs"]
        existing_item.fat_per_100g = item["fat"]
        changed += 1

    if changed:
        db.session.commit()
    return changed


def clean_ocr_number(value: str):
    text = (value or "").strip().replace(" ", "").replace(",", ".")
    if not text:
        return None
    text = text.translate(str.maketrans({
        "O": "0",
        "o": "0",
        "I": "1",
        "l": "1",
        "|": "1",
        "!": "1",
        "B": "6",
        "b": "6",
        "S": "5",
    }))
    match = re.search(r"\d+(?:\.\d+)?", text)
    return parse_decimal_text(match.group(0)) if match else None


def line_has_any(line: str, keywords: tuple[str, ...]):
    return any(keyword in line for keyword in keywords)


def extract_macro_values_from_line(line: str):
    values = []
    candidates = re.findall(r"([0-9OIl!|BS.,\s]{1,8})\s*(?:g|q|9|\u0433|\u0431)\b", line, re.IGNORECASE)
    for candidate in reversed(candidates):
        value = clean_ocr_number(candidate)
        if value is not None:
            values.append(value)
    fallback = re.search(r"([0-9OIl!|BSb.,\s]{1,8})\s*$", line, re.IGNORECASE)
    if fallback:
        value = clean_ocr_number(fallback.group(1))
        if value is not None:
            if value > 100:
                value = clean_ocr_number(str(int(value))[:-1])
            if value <= 100:
                values.append(value)
    return values


def extract_macro_value_from_line(line: str):
    values = extract_macro_values_from_line(line)
    return values[0] if values else None


def choose_ocr_value(candidates: list[float], *, require_consensus: bool = False):
    cleaned = []
    for value in candidates:
        if value is None:
            continue
        rounded = round(float(value), 1)
        cleaned.append(rounded)
    if not cleaned:
        return None
    counts = {}
    for value in cleaned:
        counts[value] = counts.get(value, 0) + 1
    best_value, best_count = max(counts.items(), key=lambda item: (item[1], -abs(item[0])))
    if require_consensus and len(counts) > 1 and best_count < 2:
        return None
    if require_consensus and len(counts) > 1 and best_count / len(cleaned) < 0.5:
        return None
    return best_value


def legacy_parse_label_text_by_rows(label_text: str):

    rows = [
        row.strip().lower()
        for row in re.split(r"[\n\r]+", label_text or "")
        if row.strip()
    ]

    candidates = {
        "calories": [],
        "protein": [],
        "carbs": [],
        "fat": [],
    }

    calorie_keywords = (
        "energy",
        "energia",
        "energjia",
        "енерг",
        "kcal",
    )

    protein_keywords = (
        "protein",
        "proteini",
        "proteina",
        "протеин",
        "белков",
    )

    carbs_keywords = (
        "carbohydrate",
        "carbs",
        "ugljeni",
        "јаглехид",
        "karbohid",
    )

    fat_keywords = (
        "fat",
        "fats",
        "mast",
        "масти",
        "yndyr",
    )

    for index, row in enumerate(rows):

        next_row = rows[index + 1] if index + 1 < len(rows) else ""
        combined = f"{row} {next_row}"

        numbers = re.findall(r"\d+(?:[.,]\d+)?", combined)
        parsed_numbers = []

        for num in numbers:
            try:
                parsed_numbers.append(float(num.replace(",", ".")))
            except:
                pass

        if not parsed_numbers:
            continue

        # calories
        if any(keyword in combined for keyword in calorie_keywords):

            kcal_matches = re.findall(
                r"(\d+(?:[.,]\d+)?)\s*kcal",
                combined,
                re.IGNORECASE,
            )

            if kcal_matches:
                try:
                    candidates["calories"].append(
                        float(kcal_matches[0].replace(",", "."))
                    )
                except:
                    pass

        # protein
        elif any(keyword in combined for keyword in protein_keywords):

            valid = [
                n for n in parsed_numbers
                if 0 < n <= 100
            ]

            if valid:
                candidates["protein"].append(max(valid))

        # carbs
        elif any(keyword in combined for keyword in carbs_keywords):

            valid = [
                n for n in parsed_numbers
                if 0 <= n <= 100
            ]

            if valid:
                candidates["carbs"].append(max(valid))

        # fat
        elif any(keyword in combined for keyword in fat_keywords):

            valid = [
                n for n in parsed_numbers
                if 0 <= n <= 100
            ]

            if valid:
                candidates["fat"].append(max(valid))

    return {
        "calories": round(candidates["calories"][0]) if candidates["calories"] else None,
        "protein": round(candidates["protein"][0], 1) if candidates["protein"] else None,
        "carbs": round(candidates["carbs"][0], 1) if candidates["carbs"] else None,
        "fat": round(candidates["fat"][0], 1) if candidates["fat"] else None,
    }

    rows = [
        row.strip().lower()
        for row in re.split(r"[\n\r]+", label_text or "")
        if row.strip()
    ]

    expanded_rows = rows[:]

    for i in range(len(rows) - 1):
        combined = f"{rows[i]} {rows[i + 1]}"
        expanded_rows.append(combined)

    rows = expanded_rows

    candidates = {
        "calories": [],
        "protein": [],
        "carbs": [],
        "fat": [],
    }

    rows = [row.strip().lower() for row in re.split(r"[\n\r]+", label_text or "") if row.strip()]
    for row in rows:
        kcal_match = re.search(r"(\d[\d\s.,]{0,7})\s*kcal", row, re.IGNORECASE)
        if kcal_match:
            candidates["calories"].append(clean_ocr_number(kcal_match.group(1)))

        if (
            line_has_any(row, ("fat", "masti", "mactu", "macru", "\u043c\u0430\u0441\u0442\u0438", "yndyr"))
            and not line_has_any(row, ("saturated", "\u0437\u0430\u0441\u0438\u0442", "ngop"))
        ):
            candidates["fat"].extend(extract_macro_values_from_line(row))

        if (
            line_has_any(row, (
                "carb",
                "karbo",
                "ugljeni",
                "jagle",
                "\u0458\u0430\u0433\u043b\u0435",
                "\u0458armex",
            ))
            and not line_has_any(row, ("sugar", "\u0448\u0435\u045c\u0435\u0440", "sheqer"))
        ):
            candidates["carbs"].extend(extract_macro_values_from_line(row))

        if (
            line_has_any(row, ("protein", "proteina", "proteini", "\u043f\u0440\u043e\u0442\u0435\u0438\u043d", "\u0440\u043e\u0442\u0435\u0438\u043d"))
        ):
            candidates["protein"].extend(extract_macro_values_from_line(row))
    return {
        "calories": choose_ocr_value(candidates["calories"]),
        "protein": choose_ocr_value(candidates["protein"], require_consensus=True),
        "carbs": choose_ocr_value(candidates["carbs"], require_consensus=True),
        "fat": choose_ocr_value(candidates["fat"], require_consensus=True),
    }


def classify_nutrition_row(row: str):
    text = (row or "").lower()
    if "nutrition facts" in text:
        return None
    if line_has_any(text, ("saturated", "saturates", "\u0437\u0430\u0441\u0438\u0442", "ngop")):
        return "saturated_fat"
    if line_has_any(text, ("sugar", "sugars", "\u0448\u0435\u045c\u0435\u0440", "sheqer")):
        return "sugars"
    if line_has_any(text, ("energy", "energia", "energjia", "\u0435\u043d\u0435\u0440\u0433", "kcal")):
        return "calories"
    if line_has_any(text, ("protein", "proteina", "proteini", "\u043f\u0440\u043e\u0442\u0435\u0438\u043d", "\u0440\u043e\u0442\u0435\u0438\u043d")):
        return "protein"
    if line_has_any(text, ("carb", "karbo", "ugljeni", "jagle", "\u0458\u0430\u0433\u043b\u0435", "\u0458armex")):
        return "carbs"
    if re.search(r"\b(?:fat|fats|masti|mactu|macru|mast\w*|yndyr\w*)\b|\u043c\u0430\u0441\u0442\u0438", text):
        return "fat"
    if line_has_any(text, ("salt", "sol", "\u0441\u043e\u043b")):
        return "salt"
    return None


def find_nutrition_label_mentions(label_text: str):
    labels = []
    for row in re.split(r"[\n\r]+", label_text or ""):
        label = classify_nutrition_row(row)
        if label:
            labels.append(label)
    return labels


def find_nutrition_value_mentions(label_text: str):
    mentions = []
    text = label_text or ""
    for match in re.finditer(r"(\d[\d\s.,]{0,7})\s*kcal", text, re.IGNORECASE):
        value = clean_ocr_number(match.group(1))
        if value is not None:
            mentions.append((match.start(), "calories", value))
    for match in re.finditer(r"(\d[\d .,]{0,7})\s*(?:g|q|9|\u0433|\u0431)\b", text, re.IGNORECASE):
        value = clean_ocr_number(match.group(1))
        prefix = text[max(0, match.start() - 8):match.start()].lower()
        if value is not None and not (round(value, 1) == 100.0 and "per" in prefix):
            mentions.append((match.start(), "grams", value))
    mentions.sort(key=lambda item: item[0])
    return mentions


def find_bare_column_gram_values(label_text: str):
    values = []
    for row in re.split(r"[\n\r]+", label_text or ""):
        stripped = row.strip()
        if not re.fullmatch(r"\d[\d .,]{0,7}", stripped):
            continue
        value = clean_ocr_number(stripped)
        if value is None or value > 100:
            continue
        if "." not in stripped and "," not in stripped and stripped.endswith("0") and value >= 10:
            value = value / 10
        values.append(value)
    return values


def add_column_order_candidates(label_text: str, candidates: dict):
    labels = find_nutrition_label_mentions(label_text)
    values = find_nutrition_value_mentions(label_text)
    if len(values) < 3 or len(labels) < 3:
        return

    gram_labels = [label for label in labels if label != "calories"]
    gram_values = [value for _, unit, value in values if unit == "grams"]
    if "calories" in labels:
        kcal_values = [value for _, unit, value in values if unit == "calories"]
        if kcal_values:
            candidates["calories"].append(kcal_values[-1])

    for label, value in zip(gram_labels, gram_values):
        if label in ("fat", "carbs", "protein") and not candidates.get(label):
            candidates[label].append(value)

    bare_values = find_bare_column_gram_values(label_text)
    for label, value in zip((label for label in labels if label in ("protein",)), bare_values):
        if not candidates.get(label):
            candidates[label].append(value)


def parse_label_text_by_rows(label_text: str):
    rows = [row.strip().lower() for row in re.split(r"[\n\r]+", label_text or "") if row.strip()]
    candidates = {
        "calories": [],
        "protein": [],
        "carbs": [],
        "fat": [],
    }

    for index, row in enumerate(rows):
        label = classify_nutrition_row(row)
        if label == "calories":
            search_text = f"{row} {rows[index + 1] if index + 1 < len(rows) else ''}"
            kcal_match = re.search(r"(\d[\d\s.,]{0,7})\s*kcal", search_text, re.IGNORECASE)
            if kcal_match:
                candidates["calories"].append(clean_ocr_number(kcal_match.group(1)))
            continue

        if label not in ("fat", "carbs", "protein"):
            continue

        values = extract_macro_values_from_line(row)
        if not values and index + 1 < len(rows) and classify_nutrition_row(rows[index + 1]) is None:
            values = extract_macro_values_from_line(rows[index + 1])
        candidates[label].extend(values)

    add_column_order_candidates(label_text, candidates)

    return {
        "calories": choose_ocr_value(candidates["calories"]),
        "protein": choose_ocr_value(candidates["protein"]),
        "carbs": choose_ocr_value(candidates["carbs"]),
        "fat": choose_ocr_value(candidates["fat"]),
    }


def parse_label_text(label_text: str):
    text = (label_text or "").lower()
    values = parse_label_text_by_rows(label_text)
    for key, patterns in LABEL_REGEX_PATTERNS.items():
        if values.get(key) is not None:
            continue
        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if not match:
                continue
            values[key] = parse_decimal_text(match.group(1))
            if values[key] is not None:
                break
        if key not in values:
            values[key] = None
    return values


def has_nutrition_label_anchor(label_text: str):
    text = (label_text or "").lower()
    return bool(
        re.search(r"\bnutrition\s+facts?\b", text)
        or re.search(r"\bper\s*100\s*g\b|\b100\s*g\b", text)
        or line_has_any(text, (
            "energy",
            "energia",
            "energjia",
            "vlera energ",
            "nutrit",
            "\u043d\u0443\u0442\u0440\u0438\u0442",
            "\u0435\u043d\u0435\u0440\u0433",
        ))
    )


def is_probable_nutrition_label(label_text: str, parsed: dict | None = None):
    parsed = parsed or parse_label_text(label_text)
    labels = find_nutrition_label_mentions(label_text)
    core_labels = [label for label in labels if label in ("calories", "fat", "carbs", "protein")]
    found_values = [
        value
        for value in (parsed or {}).values()
        if value is not None and float(value) > 0
    ]

    if not has_nutrition_label_anchor(label_text):
        return False
    if len(found_values) >= 2 and len(set(core_labels)) >= 2:
        return True
    if len(found_values) >= 1 and len(set(core_labels)) >= 3:
        return True
    return False


def build_nutrition_label_draft(label_name: str, brand: str | None, parsed: dict, ocr_text: str):
    return {
        "name": "",
        "brand": brand or "",
        "serving_label": "Scanned from label",
        "calories_per_100g": "" if parsed["calories"] is None else str(round(parsed["calories"], 1)),
        "protein_per_100g": "" if parsed["protein"] is None else str(round(parsed["protein"], 1)),
        "carbs_per_100g": "" if parsed["carbs"] is None else str(round(parsed["carbs"], 1)),
        "fat_per_100g": "" if parsed["fat"] is None else str(round(parsed["fat"], 1)),
        "ocr_text": (ocr_text or "")[:4000],
    }


def build_manual_food_draft(label_name: str, brand: str | None, ocr_text: str = ""):
    return {
        "name": "",
        "brand": brand or "",
        "serving_label": "",
        "calories_per_100g": "",
        "protein_per_100g": "",
        "carbs_per_100g": "",
        "fat_per_100g": "",
        "ocr_text": (ocr_text or "")[:4000],
    }


def score_label_text_candidates(label_text: str):
    parsed = parse_label_text(label_text)
    found_values = sum(1 for value in parsed.values() if value is not None)
    digit_count = len(re.findall(r"\d", label_text or ""))
    return found_values, digit_count, -len(label_text or "")


def build_label_ocr_images(image):
    base = ImageOps.exif_transpose(image)
    variants = [base]
    for angle in (-2, 2):
        variants.append(base.rotate(angle, expand=True, fillcolor="white"))

    processed = []
    for variant in variants:
        grayscale = ImageOps.grayscale(variant)
        autocontrast = ImageOps.autocontrast(grayscale)
        scale = 2 if max(autocontrast.size) < 1600 else 1
        enlarged = autocontrast.resize(
            (autocontrast.width * scale, autocontrast.height * scale),
            Image.Resampling.LANCZOS,
        ).filter(ImageFilter.UnsharpMask(radius=1, percent=150, threshold=3))
        threshold = enlarged.point(lambda pixel: 255 if pixel > 160 else 0)
        processed.extend([enlarged, threshold])
    return processed


def get_google_vision_language_hints():
    return [
        hint.strip()
        for hint in app.config.get("GOOGLE_VISION_LANGUAGE_HINTS", "").split(",")
        if hint.strip()
    ]


def scan_label_with_google_vision(image_path: str):
    api_key = app.config.get("GOOGLE_VISION_API_KEY", "").strip()
    if not api_key:
        raise RuntimeError("Google Vision API key is not configured.")

    with open(image_path, "rb") as fh:
        encoded_image = base64.b64encode(fh.read()).decode("ascii")

    feature_type = app.config.get("GOOGLE_VISION_FEATURE_TYPE", "DOCUMENT_TEXT_DETECTION").strip() or "DOCUMENT_TEXT_DETECTION"
    request_payload = {
        "requests": [{
            "image": {"content": encoded_image},
            "features": [{"type": feature_type}],
        }]
    }
    language_hints = get_google_vision_language_hints()
    if language_hints:
        request_payload["requests"][0]["imageContext"] = {"languageHints": language_hints}

    endpoint = f"https://vision.googleapis.com/v1/images:annotate?{urlencode({'key': api_key})}"
    req = Request(
        endpoint,
        data=json.dumps(request_payload).encode("utf-8"),
        headers={"Content-Type": "application/json"},
        method="POST",
    )

    try:
        with urlopen(req, timeout=20) as resp:
            payload = json.loads(resp.read().decode("utf-8"))
    except (HTTPError, URLError, TimeoutError, json.JSONDecodeError) as exc:
        raise RuntimeError("Google Vision could not read that image right now.") from exc

    if payload.get("error"):
        message = payload["error"].get("message") or "Google Vision returned an error."
        raise RuntimeError(message)

    responses = payload.get("responses") or []
    if not responses:
        return ""
    response = responses[0]
    if response.get("error"):
        message = response["error"].get("message") or "Google Vision returned an error."
        raise RuntimeError(message)

    full_text = response.get("fullTextAnnotation", {}).get("text")
    if full_text:
        return full_text.strip()

    annotations = response.get("textAnnotations") or []
    if annotations and annotations[0].get("description"):
        return annotations[0]["description"].strip()
    return ""


def scan_label_image_text(image):
    lang = get_tesseract_languages()
    base_config = "--oem 1"
    configs = (f"{base_config} --psm 6", f"{base_config} --psm 4")

    best_text = ""
    best_score = (-1, -1, 0)
    for ocr_image in build_label_ocr_images(image):
        for config in configs:
            text = pytesseract.image_to_string(ocr_image, lang=lang, config=config)
            if not text.strip():
                continue
            score = score_label_text_candidates(text)
            if score > best_score:
                best_score = score
                best_text = text
            if score[0] == 4:
                return text.strip()

    return best_text.strip()


def scan_label_file_text(image_path: str):
    if google_vision_enabled():
        try:
            return scan_label_with_google_vision(image_path)
        except Exception:
            # A configured key doesn't mean a working one (expired, over
            # quota, revoked...). Fall back to local Tesseract rather than
            # taking label scanning down entirely whenever that happens.
            if not tesseract_scan_enabled():
                raise

    with Image.open(image_path) as image:
        return scan_label_image_text(image)


def build_nutrition_summary(logs: list[FoodLogEntry], calorie_target: int | None):
    totals = {
        "calories": round(sum(log.calories for log in logs), 1),
        "protein": round(sum(log.protein for log in logs), 1),
        "carbs": round(sum(log.carbs for log in logs), 1),
        "fat": round(sum(log.fat for log in logs), 1),
    }
    meal_sections = []
    for meal_type in MEAL_TYPES:
        meal_logs = [log for log in logs if log.meal_type == meal_type]
        if not meal_logs:
            continue
        meal_sections.append({
            "key": meal_type,
            "label": meal_type_label(meal_type),
            "items": meal_logs,
            "calories": round(sum(log.calories for log in meal_logs), 1),
        })

    goal = calorie_target or 0
    # Calculate macro goals based on calorie target
    protein_goal = round((goal * 0.35) / 4) if goal > 0 else 0
    carbs_goal = round((goal * 0.40) / 4) if goal > 0 else 0
    fat_goal = round((goal * 0.25) / 9) if goal > 0 else 0

    if goal > 0:
        ratio = totals["calories"] / goal
        percent = int(round(ratio * 100))
        ring_percent = max(0, min(percent, 100))
        remaining = round(goal - totals["calories"], 1)
        if ratio < 0.85:
            tone = "safe"
            status = f"{remaining:.0f} kcal left"
        elif ratio <= 1:
            tone = "near"
            status = f"{remaining:.0f} kcal left"
        else:
            tone = "over"
            status = f"{abs(remaining):.0f} kcal over"
    else:
        percent = 0
        ring_percent = 0
        tone = "safe"
        status = "Set a daily target"

    return {
        "totals": totals,
        "meal_sections": meal_sections,
        "goal": goal,
        "protein_goal": protein_goal,
        "carbs_goal": carbs_goal,
        "fat_goal": fat_goal,
        "percent": percent,
        "ring_percent": ring_percent,
        "tone": tone,
        "status": status,
        "has_logs": bool(logs),
    }


def build_strength_summary(entries: list[StrengthLogEntry], selected_exercise: str | None):
    exercises = sorted({entry.exercise.strip() for entry in entries if (entry.exercise or "").strip()})
    selected = (selected_exercise or "").strip()
    if selected and selected not in exercises:
        selected = ""
    if not selected and exercises:
        counts = {}
        for exercise in exercises:
            counts[exercise] = sum(1 for entry in entries if entry.exercise == exercise)
        selected = max(exercises, key=lambda exercise: (counts[exercise], exercise.lower()))

    filtered = [entry for entry in entries if entry.exercise == selected] if selected else []
    grouped = {}
    for entry in filtered:
        day = entry.logged_at.date()
        grouped.setdefault(day, []).append(entry)

    progress_points = []
    for day in sorted(grouped):
        usable = [entry for entry in grouped[day] if entry.weight is not None]
        if not usable:
            continue
        top = max(usable, key=lambda entry: ((entry.weight or 0), (entry.reps or 0), entry.id))
        progress_points.append({
            "date": day.strftime("%Y-%m-%d"),
            "weight": round(top.weight or 0, 1),
            "reps": top.reps or 0,
            "e1rm": estimate_one_rep_max(top.weight, top.reps),
        })

    best_weight_entry = None
    best_volume_entry = None
    for entry in filtered:
        if entry.weight is not None and (best_weight_entry is None or (entry.weight, entry.reps or 0) > (best_weight_entry.weight, best_weight_entry.reps or 0)):
            best_weight_entry = entry
        if entry.weight is None or entry.reps is None:
            continue
        if best_volume_entry is None or ((entry.weight * entry.reps), entry.weight) > ((best_volume_entry.weight * best_volume_entry.reps), best_volume_entry.weight):
            best_volume_entry = entry

    latest = filtered[-1] if filtered else None
    return {
        "exercises": exercises,
        "selected_exercise": selected,
        "entries": filtered,
        "points": progress_points,
        "total_sets": len(filtered),
        "workout_count": len({entry.logged_at.date() for entry in filtered}),
        "best_weight_entry": best_weight_entry,
        "best_volume_entry": best_volume_entry,
        "latest_entry": latest,
    }


def parse_strength_date_range(start_raw: str | None, end_raw: str | None, entries: list[StrengthLogEntry]):
    start_date = parse_iso_date(start_raw)
    end_date = parse_iso_date(end_raw)
    if entries:
        min_day = entries[0].logged_at.date()
        max_day = entries[-1].logged_at.date()
    else:
        min_day = date.today()
        max_day = date.today()

    if end_date is None:
        end_date = max_day
    if start_date is None:
        start_date = min_day
    if start_date > end_date:
        start_date, end_date = end_date, start_date
    return start_date, end_date


def get_latest_weight_on_or_before(points: list[Measurement], day: date):
    candidates = [point for point in points if point.date.date() <= day and point.weight is not None]
    return candidates[-1].weight if candidates else None


def build_strength_poster_data(
    entries: list[StrengthLogEntry],
    weight_points: list[Measurement],
    start_date: date,
    end_date: date,
    manual_lifts: list[str],
):
    ranged = [entry for entry in entries if start_date <= entry.logged_at.date() <= end_date and entry.weight is not None]
    grouped = {}
    for entry in ranged:
        grouped.setdefault(entry.exercise, []).append(entry)

    progress_rows = []
    for exercise, exercise_entries in grouped.items():
        sorted_entries = sorted(exercise_entries, key=lambda entry: (entry.logged_at, entry.id))
        weakest_entry = min(
            sorted_entries,
            key=lambda entry: (
                entry.weight if entry.weight is not None else float("inf"),
                entry.reps if entry.reps is not None else float("inf"),
                entry.logged_at,
            ),
        )
        strongest_entry = max(
            sorted_entries,
            key=lambda entry: (
                entry.weight if entry.weight is not None else float("-inf"),
                entry.reps if entry.reps is not None else float("-inf"),
                entry.logged_at,
            ),
        )
        start_weight = round(weakest_entry.weight or 0, 1)
        end_weight = round(strongest_entry.weight or 0, 1)
        delta = round(end_weight - start_weight, 1)
        pct = round((delta / start_weight) * 100, 1) if start_weight > 0 else 0.0
        progress_rows.append({
            "exercise": exercise,
            "start": start_weight,
            "end": end_weight,
            "delta": delta,
            "pct": pct,
            "start_reps": weakest_entry.reps or 0,
            "end_reps": strongest_entry.reps or 0,
            "start_date": weakest_entry.logged_at.date(),
            "end_date": strongest_entry.logged_at.date(),
        })

    selected_rows = []
    chosen_names = [name.strip() for name in manual_lifts if (name or "").strip()]
    if chosen_names:
        by_name = {row["exercise"]: row for row in progress_rows}
        for name in chosen_names:
            row = by_name.get(name)
            if row:
                selected_rows.append(row)
        selected_rows = selected_rows[:6]
    else:
        selected_rows = sorted(
            progress_rows,
            key=lambda row: (row["delta"], row["pct"], row["end"]),
            reverse=True,
        )[:6]

    before_weight = weight_points[0].weight if weight_points else None
    after_weight = weight_points[-1].weight if weight_points else None
    body_delta = round(after_weight - before_weight, 1) if before_weight is not None and after_weight is not None else None
    weeks = max(1, ceil(((end_date - start_date).days + 1) / 7))

    return {
        "start_date": start_date,
        "end_date": end_date,
        "weeks": weeks,
        "before_weight": before_weight,
        "after_weight": after_weight,
        "body_delta": body_delta,
        "rows": selected_rows,
        "available_exercises": sorted(grouped.keys()),
    }


_poster_font_cache = {}


def poster_font(size: int, bold: bool = False):
    # Bundled locally (static/fonts/Manrope-Variable.ttf) rather than resolved from
    # OS/venv font paths — those never resolve inside the Railway/Docker container,
    # which silently fell back to PIL's tiny default bitmap font in production.
    key = (size, bold)
    if key not in _poster_font_cache:
        root = os.path.dirname(os.path.abspath(__file__))
        font_path = os.path.join(root, "static", "fonts", "Manrope-Variable.ttf")
        font = ImageFont.truetype(font_path, size=size)
        font.set_variation_by_axes([800 if bold else 500])
        _poster_font_cache[key] = font
    return _poster_font_cache[key]


def draw_text_fit(draw, text: str, xy, font, fill, max_width: int, max_lines: int = 2, line_gap: int = 6):
    words = str(text).split()
    lines = []
    current = ""
    for word in words:
        trial = f"{current} {word}".strip()
        bbox = draw.textbbox((0, 0), trial, font=font)
        if bbox[2] - bbox[0] <= max_width or not current:
            current = trial
        else:
            lines.append(current)
            current = word
        if len(lines) >= max_lines:
            break
    if current and len(lines) < max_lines:
        lines.append(current)
    if words and len(lines) == max_lines and " ".join(lines).split() != words:
        while lines[-1] and draw.textbbox((0, 0), lines[-1] + "...", font=font)[2] > max_width:
            lines[-1] = lines[-1][:-1].rstrip()
        lines[-1] = lines[-1] + "..."
    x, y = xy
    line_height = draw.textbbox((0, 0), "Ag", font=font)[3] + line_gap
    for line in lines:
        draw.text((x, y), line, fill=fill, font=font)
        y += line_height
    return y


def font_that_fits(text: str, size: int, max_width: int, bold: bool = False, min_size: int = 24):
    text = str(text)
    for candidate_size in range(size, min_size - 1, -2):
        font = poster_font(candidate_size, bold)
        bbox = ImageDraw.Draw(Image.new("RGB", (1, 1))).textbbox((0, 0), text, font=font)
        if bbox[2] - bbox[0] <= max_width:
            return font
    return poster_font(min_size, bold)


def render_strength_poster_png(client_name: str, poster: dict):
    width, height = 1080, 1350
    img = Image.new("RGB", (width, height), "#111315")
    draw = ImageDraw.Draw(img, "RGBA")

    def text_width(text, font):
        bbox = draw.textbbox((0, 0), str(text), font=font)
        return bbox[2] - bbox[0]

    def right_text(x_right, y, text, font, fill):
        draw.text((x_right - text_width(text, font), y), text, fill=fill, font=font)

    def rounded_box(box, fill, outline=None, radius=8, width=1):
        draw.rounded_rectangle(box, radius=radius, fill=fill, outline=outline, width=width)

    for y in range(height):
        ratio = y / height
        r = int(15 + (22 - 15) * ratio)
        g = int(17 + (22 - 17) * ratio)
        b = int(18 + (20 - 18) * ratio)
        draw.line((0, y, width, y), fill=(r, g, b, 255))

    for x in range(72, width, 72):
        draw.line((x, 0, x, height), fill=(255, 255, 255, 9), width=1)
    for y in range(80, height, 80):
        draw.line((0, y, width, y), fill=(255, 255, 255, 7), width=1)

    draw.rectangle((0, 0, 18, height), fill=(221, 255, 80, 255))
    draw.polygon([(840, 0), (1080, 0), (1080, 390), (936, 338)], fill=(221, 255, 80, 18))
    draw.polygon([(0, 1040), (360, 1350), (0, 1350)], fill=(221, 255, 80, 13))

    section_font = poster_font(48, True)
    label_font = poster_font(18, True)
    body_font = poster_font(23)
    small_font = poster_font(20)
    tiny_font = poster_font(15, True)
    lift_font = poster_font(27, True)
    delta_font = poster_font(42, True)

    ink = "#f4f1ea"
    muted = "#a8adb2"
    soft = "#d4d8da"
    accent = "#ddff50"
    line = (244, 241, 234, 46)
    panel_dark = (24, 28, 30, 238)
    panel_lift = (29, 34, 36, 246)
    date_range = f"{poster['start_date'].strftime('%d %b %Y')} - {poster['end_date'].strftime('%d %b %Y')}"
    client_title = str(client_name or "Client").strip()
    fitted_title = font_that_fits(client_title, 82, 660, bold=True, min_size=42)
    draw.text((74, 92), client_title, fill=ink, font=fitted_title)
    draw.text((78, 194), date_range, fill=muted, font=body_font)

    weeks_label = f"{poster['weeks']} weeks"
    weeks_font = font_that_fits(weeks_label, 54, 260, bold=True, min_size=34)
    right_text(1008, 94, weeks_label, weeks_font, accent)
    draw.line((74, 250, 1006, 250), fill=line, width=2)

    summary_items = [
        ("BODY START", f"{poster['before_weight']:.1f} kg" if poster["before_weight"] is not None else "-"),
        ("BODY END", f"{poster['after_weight']:.1f} kg" if poster["after_weight"] is not None else "-"),
        ("CHANGE", f"{poster['body_delta']:+.1f} kg" if poster["body_delta"] is not None else "-"),
    ]
    has_body_summary = any(value != "-" for _label, value in summary_items)
    if has_body_summary:
        summary_y = 288
        summary_cols = [(74, 300), (390, 616), (706, 1006)]
        for idx, (label, value) in enumerate(summary_items):
            x1, x2 = summary_cols[idx]
            draw.line((x1, summary_y, x2, summary_y), fill=line, width=2)
            draw.text((x1, summary_y + 18), label, fill=muted, font=tiny_font)
            value_font = font_that_fits(value, 44, x2 - x1 - 4, bold=True, min_size=28)
            draw.text((x1, summary_y + 48), value, fill=ink if label != "CHANGE" else accent, font=value_font)
        lift_title_y = 430
    else:
        lift_title_y = 300

    draw.text((74, lift_title_y), "LIFT CHANGES", fill=ink, font=section_font)

    rows = poster["rows"][:6]
    y = lift_title_y + 104
    row_h = 112
    for index, row in enumerate(rows, start=1):
        row_fill = panel_lift if index % 2 else panel_dark
        row_outline = (255, 255, 255, 10)
        rounded_box((74, y, 1006, y + row_h), row_fill, row_outline, radius=8, width=1)
        row_ink = ink
        row_muted = muted
        row_soft = soft
        delta_fill = accent

        draw.text((100, y + 30), f"{index:02d}", fill=accent, font=label_font)
        draw.line((178, y + 22, 178, y + row_h - 22), fill=line, width=2)

        title_font = font_that_fits(row["exercise"], 28, 520, bold=True, min_size=20)
        draw_text_fit(draw, row["exercise"], (202, y + 14), title_font, row_ink, 520, max_lines=1)

        start_label = f"{row['start']:.1f} kg x {row['start_reps']}"
        end_label = f"{row['end']:.1f} kg x {row['end_reps']}"
        compare_y = y + 69
        draw.text((202, compare_y), "BEFORE", fill=row_soft, font=tiny_font)
        start_font = font_that_fits(start_label, 22, 164, bold=True, min_size=17)
        draw.text((202, compare_y + 17), start_label, fill=row_muted, font=start_font)

        arrow_fill = (221, 255, 80, 150)
        draw.line((388, compare_y + 24, 424, compare_y + 24), fill=arrow_fill, width=3)
        draw.polygon(
            [(424, compare_y + 24), (415, compare_y + 17), (415, compare_y + 31)],
            fill=arrow_fill,
        )

        draw.text((450, compare_y), "AFTER", fill=accent, font=tiny_font)
        end_font = font_that_fits(end_label, 22, 180, bold=True, min_size=17)
        draw.text((450, compare_y + 17), end_label, fill=row_ink, font=end_font)

        delta_label = f"{row['delta']:+.1f} kg"
        pct_label = f"{row['pct']:+.1f}%"
        right_text(970, y + 16, delta_label, delta_font, delta_fill)
        pct_font = font_that_fits(pct_label, 20, 120, bold=False, min_size=17)
        pct_w = text_width(pct_label, pct_font)
        rounded_box((970 - pct_w - 24, y + 58, 970, y + 84), (244, 241, 234, 18), None, radius=13)
        right_text(958, y + 62, pct_label, pct_font, row_muted)
        y += row_h + 10

    if not rows:
        rounded_box((74, y, 1006, y + 150), panel_dark, None, radius=8)
        draw.text((110, y + 56), "No comparable lift data in this range.", fill=accent, font=body_font)

    output = io.BytesIO()
    img.save(output, format="PNG")
    output.seek(0)
    return output


def normalize_strength_header(value):
    return str(value or "").strip()


def read_strength_csv_rows(uploaded_file):
    raw = uploaded_file.stream.read()
    uploaded_file.stream.seek(0)
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("latin-1")
    return list(csv.DictReader(io.StringIO(text)))


def read_strength_workbook_rows(uploaded_file):
    if openpyxl is None:
        raise ValueError("Excel support is not installed.")
    uploaded_file.stream.seek(0)
    workbook = openpyxl.load_workbook(uploaded_file.stream, read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        rows = sheet.iter_rows(values_only=True)
        headers = [normalize_strength_header(value) for value in next(rows, [])]
        parsed = []
        for values in rows:
            parsed.append({
                headers[index]: value
                for index, value in enumerate(values)
                if index < len(headers) and headers[index]
            })
        return parsed
    finally:
        uploaded_file.stream.seek(0)
        workbook.close()


def strength_upload_is_workbook(uploaded_file, filename: str):
    extension = os.path.splitext(filename.lower())[1]
    uploaded_file.stream.seek(0)
    signature = uploaded_file.stream.read(4)
    uploaded_file.stream.seek(0)
    return signature == b"PK\x03\x04" or extension in (".xlsx", ".xlsm", ".xltx", ".xltm")


def detect_strength_export_source(rows: list[dict]):
    if not rows:
        return None
    keys = {str(key or "").strip().lower() for key in rows[0].keys()}
    if {"title", "date", "exercise", "weight", "reps"}.issubset(keys):
        return "lyfta"
    if {"title", "start_time", "exercise_title", "weight_kg", "reps"}.issubset(keys):
        return "hevy"
    return None


def get_row_value(row: dict, *keys):
    lowered = {str(key or "").strip().lower(): value for key, value in row.items()}
    for key in keys:
        if key.lower() in lowered:
            return lowered[key.lower()]
    return None


def import_strength_export(client_id: int, uploaded_file, source_name: str):
    rows = (
        read_strength_workbook_rows(uploaded_file)
        if strength_upload_is_workbook(uploaded_file, source_name)
        else read_strength_csv_rows(uploaded_file)
    )
    if not rows:
        return {"created": 0, "skipped": 0, "source": "unknown"}

    source = detect_strength_export_source(rows)
    if source is None:
        raise ValueError("Unsupported strength export format.")

    StrengthLogEntry.query.filter_by(client_id=client_id).delete()

    created = 0
    skipped = 0
    for row in rows:
        if source == "hevy":
            exercise = str(get_row_value(row, "exercise_title") or "").strip()
            logged_at = parse_strength_export_datetime(get_row_value(row, "start_time"))
            weight = parse_decimal_text(get_row_value(row, "weight_kg"))
            reps_value = parse_decimal_text(get_row_value(row, "reps"))
            workout_title = str(get_row_value(row, "title") or "").strip() or None
            rir_rpe = str(get_row_value(row, "rpe") or "").strip() or None
            duration = str(get_row_value(row, "duration_seconds") or "").strip() or None
            set_type = str(get_row_value(row, "set_type") or "").strip() or None
        else:
            exercise = str(get_row_value(row, "Exercise") or "").strip()
            logged_at = parse_strength_export_datetime(get_row_value(row, "Date"))
            weight = parse_decimal_text(get_row_value(row, "Weight"))
            reps_value = parse_decimal_text(get_row_value(row, "Reps"))
            workout_title = str(get_row_value(row, "Title") or "").strip() or None
            rir_rpe = str(get_row_value(row, "RIR/RPE") or "").strip() or None
            duration = str(get_row_value(row, "Duration") or "").strip() or None
            set_type = str(get_row_value(row, "Set Type") or "").strip() or None

        if not exercise or not logged_at:
            skipped += 1
            continue

        reps = int(reps_value) if reps_value is not None else None
        entry = StrengthLogEntry(
            client_id=client_id,
            source=source,
            workout_title=workout_title,
            logged_at=logged_at,
            exercise=exercise[:160],
            weight=weight,
            reps=reps,
            rir_rpe=rir_rpe,
            duration=duration,
            set_type=set_type,
            source_file=source_name[:255],
        )
        db.session.add(entry)
        created += 1

    db.session.commit()
    return {"created": created, "skipped": skipped, "source": source}


def serialize_food_log(log: FoodLogEntry):
    return {
        "id": log.id,
        "food_name": log.food_name,
        "quantity_grams": log.quantity_grams,
        "calories": log.calories,
        "protein": log.protein,
        "carbs": log.carbs,
        "fat": log.fat,
        "note": log.note or "",
    }


def serialize_meal_sections(meal_sections: list[dict]):
    return [
        {
            "key": meal["key"],
            "label": meal["label"],
            "calories": meal["calories"],
            "items": [serialize_food_log(item) for item in meal["items"]],
        }
        for meal in meal_sections
    ]


def nutrition_summary_payload(nutrition_summary: dict):
    return {
        "totals": nutrition_summary["totals"],
        "goal": nutrition_summary["goal"],
        "protein_goal": nutrition_summary["protein_goal"],
        "carbs_goal": nutrition_summary["carbs_goal"],
        "fat_goal": nutrition_summary["fat_goal"],
        "percent": nutrition_summary["percent"],
        "ring_percent": nutrition_summary["ring_percent"],
        "tone": nutrition_summary["tone"],
        "status": nutrition_summary["status"],
        "meal_sections": serialize_meal_sections(nutrition_summary["meal_sections"]),
        "has_logs": nutrition_summary["has_logs"],
    }


def nutrition_day_label(day: date):
    today = date.today()
    if day == today:
        return "Today"
    if day == today - timedelta(days=1):
        return "Yesterday"
    return day.strftime("%A")


def allowed_photo_file(filename: str):
    name = (filename or "").lower()
    return name.endswith(".jpg") or name.endswith(".jpeg") or name.endswith(".png") or name.endswith(".webp")


def validate_uploaded_image(file_storage, *, max_bytes: int | None = None):
    if not file_storage or not file_storage.filename:
        return "Please choose a photo file."
    if not allowed_photo_file(file_storage.filename):
        return "Allowed formats: .jpg, .jpeg, .png, .webp"

    limit = max_bytes or app.config["MAX_PHOTO_UPLOAD_BYTES"]
    stream = file_storage.stream
    try:
        stream.seek(0, os.SEEK_END)
        size = stream.tell()
        stream.seek(0)
    except (OSError, AttributeError):
        size = file_storage.content_length or 0

    if size and size > limit:
        stream.seek(0)
        return f"Photo is too large. Max size is {limit // (1024 * 1024)} MB."

    try:
        with Image.open(stream) as img:
            img.verify()
    except Exception:
        stream.seek(0)
        return "That file does not look like a valid image."

    stream.seek(0)
    return None


def parse_phone(value):
    phone = (value or "").strip()
    if phone == "":
        return ""
    if not re.fullmatch(r"\+?\d+", phone):
        return None
    return phone


def build_session_calendar(sessions: list[SessionLog], target_date: date | None = None):
    target = target_date or date.today()
    year = target.year
    month = target.month
    month_name = calendar.month_name[month]

    first_weekday, days_in_month = calendar.monthrange(year, month)  # Mon=0
    offset = first_weekday

    session_counts = {}
    for s in sessions:
        d = s.date.date()
        if d.year == year and d.month == month:
            session_counts[d.day] = session_counts.get(d.day, 0) + 1

    cells = []
    for _ in range(offset):
        cells.append({"day": None, "count": 0, "is_today": False})
    for day_num in range(1, days_in_month + 1):
        today = date.today()
        cells.append({
            "day": day_num,
            "count": session_counts.get(day_num, 0),
            "is_today": today.year == year and today.month == month and today.day == day_num,
        })
    while len(cells) % 7 != 0:
        cells.append({"day": None, "count": 0, "is_today": False})

    return {
        "year": year,
        "month": month,
        "month_name": month_name,
        "weekday_labels": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"],
        "cells": cells,
    }


def week_start(d: date) -> date:
    # Monday start
    return d - timedelta(days=d.weekday())


def add_months(d: date, months: int) -> date:
    # add months keeping day if possible (simple safe method)
    year = d.year + (d.month - 1 + months) // 12
    month = (d.month - 1 + months) % 12 + 1
    day = d.day

    # clamp day to month length
    # month lengths with leap year check
    def days_in_month(y, m):
        if m in (1, 3, 5, 7, 8, 10, 12):
            return 31
        if m in (4, 6, 9, 11):
            return 30
        # feb
        leap = (y % 4 == 0 and y % 100 != 0) or (y % 400 == 0)
        return 29 if leap else 28

    dim = days_in_month(year, month)
    if day > dim:
        day = dim

    return date(year, month, day)


def payment_due_date(payment: Payment) -> date:
    if payment.due_date_override:
        return payment.due_date_override
    return add_months(payment.start_date, payment.months)


BUDGET_DAY_NAMES = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]


def get_budget_week(ws: date):
    return BudgetWeek.query.filter_by(week_start=ws).first()


def build_budget_week_view(ws: date):
    week = get_budget_week(ws)
    weekly_budget = week.weekly_budget if week else 0
    daily_budget = round(weekly_budget / 7) if weekly_budget else 0

    entries = {
        e.day: e for e in BudgetDay.query.filter(
            BudgetDay.day >= ws, BudgetDay.day <= ws + timedelta(days=6)
        ).all()
    }

    today = date.today()
    days = []
    total_spent = 0
    for i in range(7):
        d = ws + timedelta(days=i)
        entry = entries.get(d)
        logged = entry is not None
        spent = entry.spent if entry else 0
        if logged:
            total_spent += spent
        over = logged and daily_budget > 0 and spent > daily_budget
        pct = min(100, round((spent / daily_budget) * 100)) if daily_budget > 0 else 0
        if not logged:
            tone = "empty"
        elif over:
            tone = "over"
        elif daily_budget > 0 and spent >= daily_budget * 0.85:
            tone = "warn"
        else:
            tone = "ok"
        days.append({
            "date": d,
            "name": BUDGET_DAY_NAMES[i],
            "spent": spent,
            "logged": logged,
            "leftover": daily_budget - spent,
            "over": over,
            "pct": pct,
            "tone": tone,
            "is_today": d == today,
            "is_future": d > today,
        })

    weekly_remaining = weekly_budget - total_spent

    return {
        "week_start": ws,
        "week_end": ws + timedelta(days=6),
        "weekly_budget": weekly_budget,
        "daily_budget": daily_budget,
        "days": days,
        "total_spent": total_spent,
        "weekly_remaining": weekly_remaining,
        "week_over": weekly_budget > 0 and total_spent > weekly_budget,
        "has_budget": weekly_budget > 0,
        "over_amount": max(0, total_spent - weekly_budget),
    }


def budget_day_ajax_payload(d: date):
    view = build_budget_week_view(week_start(d))
    day_data = next(x for x in view["days"] if x["date"] == d)
    return {
        "ok": True,
        "day": {
            "date": day_data["date"].isoformat(),
            "spent": day_data["spent"],
            "logged": day_data["logged"],
            "leftover": day_data["leftover"],
            "over": day_data["over"],
            "pct": day_data["pct"],
            "tone": day_data["tone"],
            "has_budget": view["has_budget"],
        },
        "week": {
            "total_spent": view["total_spent"],
            "weekly_remaining": view["weekly_remaining"],
            "week_over": view["week_over"],
            "over_amount": view["over_amount"],
        },
        "all_time_savings": compute_all_time_savings(),
    }


def compute_all_time_savings() -> int:
    total = 0
    week_daily_cache = {}
    for entry in BudgetDay.query.all():
        ws = week_start(entry.day)
        if ws not in week_daily_cache:
            week = get_budget_week(ws)
            week_daily_cache[ws] = round(week.weekly_budget / 7) if week else 0
        total += week_daily_cache[ws] - entry.spent
    return total


def seed_admin():
    existing = User.query.filter_by(username="admin").first()
    if not existing:
        default_admin_username = os.environ.get("ADMIN_DEFAULT_USERNAME", "admin")
        default_admin_password = os.environ.get("ADMIN_DEFAULT_PASSWORD", "admin123")
        u = User(
            username=default_admin_username,
            password_hash=generate_password_hash(default_admin_password),
            role="admin",
            client_id=None,
        )
        db.session.add(u)
        db.session.commit()
        print(f"Created default admin: {default_admin_username} / {default_admin_password}")


@app.cli.command("seed-admin")
def seed_admin_command():
    """Create default admin user if it doesn't exist."""
    seed_admin()
    click.echo("seed-admin complete")


@app.cli.command("import-capnutra")
def import_capnutra_command():
    """Import CAPNUTRA Serbian FCDB foods into the shared nutrition library."""

    def progress(field, page, pages, total):
        click.echo(f"{field}: page {page}/{pages}, collected {total} foods")

    stats = import_capnutra_foods(fetch_capnutra_food_library(progress=progress), prune_missing=True)
    click.echo(
        "CAPNUTRA import complete: "
        f"{stats['created']} created, {stats['updated']} updated, "
        f"{stats['removed']} removed, {stats['total']} active."
    )


def get_current_plan(client_id: int):
    """Returns (sessions_per_week, current_status dict or None) based on latest payment."""
    today = date.today()

    latest_payment = (
        Payment.query.filter_by(client_id=client_id)
        .order_by(Payment.start_date.desc(), Payment.paid_on.desc(), Payment.id.desc())
        .first()
    )

    if not latest_payment:
        return None, None

    paid_until = payment_due_date(latest_payment)
    days_left = (paid_until - today).days

    status = {
        "amount_paid": latest_payment.amount_paid,
        "start_date": latest_payment.start_date,
        "months": latest_payment.months,
        "paid_until": paid_until,
        "days_left": days_left,
        "sessions_per_week": latest_payment.sessions_per_week,
    }
    return latest_payment.sessions_per_week, status


def compute_sessions(client: Client, sessions_per_week: int):
    """Compute used/remaining for current week including rollover."""
    today = date.today()
    ws = week_start(today)

    used_this_week = (
        SessionLog.query.filter_by(client_id=client.id)
        .filter(SessionLog.date >= datetime(ws.year, ws.month, ws.day, tzinfo=timezone.utc))
        .count()
    )

    bonus = 0
    if client.rollover_for_week == ws and (client.rollover_bonus or 0) > 0:
        bonus = client.rollover_bonus

    allowed = max((sessions_per_week or 0) + bonus, 0)
    remaining = max(allowed - used_this_week, 0)

    return used_this_week, remaining, bonus, allowed


def csv_download_response(filename: str, headers: list[str], rows: list[list]):
    output = io.StringIO()
    output.write("\ufeff")  # UTF-8 BOM for spreadsheet compatibility.
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerows(rows)
    return Response(
        output.getvalue(),
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def resolve_db_file_path():
    db_path = db.engine.url.database
    if not db_path:
        return None

    candidates = []
    if os.path.isabs(db_path):
        candidates.append(db_path)
    else:
        candidates.append(os.path.join(app.instance_path, db_path))
        candidates.append(os.path.join(app.root_path, db_path))
        candidates.append(os.path.abspath(db_path))

    for path in candidates:
        if os.path.exists(path):
            return path
    return None


def get_or_404(model, object_id):
    obj = db.session.get(model, object_id)
    if obj is None:
        abort(404)
    return obj


def login_throttle_status(keys: list[str]):
    now = utc_now()
    max_seconds_left = 0
    window_seconds = app.config["LOGIN_WINDOW_SECONDS"]

    for key in keys:
        state = LoginThrottle.query.filter_by(key=key).first()
        if not state:
            continue

        lock_until = state.lock_until
        if lock_until and lock_until.tzinfo is not None:
            lock_until = lock_until.astimezone(timezone.utc).replace(tzinfo=None)
        if lock_until and lock_until > now:
            seconds_left = ceil((lock_until - now).total_seconds())
            if seconds_left > max_seconds_left:
                max_seconds_left = seconds_left
            continue

        first_ts = state.first_ts
        if first_ts.tzinfo is not None:
            first_ts = first_ts.astimezone(timezone.utc).replace(tzinfo=None)
        if (now - first_ts).total_seconds() > window_seconds:
            db.session.delete(state)

    if db.session.deleted:
        db.session.commit()

    return max_seconds_left > 0, max_seconds_left


def login_throttle_failed(keys: list[str]):
    now = utc_now()
    window_seconds = app.config["LOGIN_WINDOW_SECONDS"]
    max_attempts = app.config["LOGIN_MAX_ATTEMPTS"]
    lock_seconds = app.config["LOGIN_LOCK_SECONDS"]

    for key in keys:
        state = LoginThrottle.query.filter_by(key=key).first()
        if not state:
            state = LoginThrottle(key=key, count=0, first_ts=now, lock_until=None)
            db.session.add(state)
        else:
            first_ts = state.first_ts
            if first_ts.tzinfo is not None:
                first_ts = first_ts.astimezone(timezone.utc).replace(tzinfo=None)
            if (now - first_ts).total_seconds() > window_seconds:
                state.count = 0
                state.first_ts = now
                state.lock_until = None

        state.count += 1
        if state.count >= max_attempts:
            state.lock_until = now + timedelta(seconds=lock_seconds)

    db.session.commit()


def login_throttle_success(keys: list[str]):
    states = LoginThrottle.query.filter(LoginThrottle.key.in_(keys)).all()
    if not states:
        return
    for state in states:
        db.session.delete(state)
    db.session.commit()


# =========================
# Auth Routes
# =========================
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "GET":
        uid = session.get("user_id")
        if uid:
            user = db.session.get(User, uid)
            if user and user.role != "disabled":
                if user.role == "admin":
                    return redirect(url_for("index"))
                if user.client_id:
                    return redirect(url_for("client_profile", client_id=user.client_id, tab="info"))
        return render_template("login.html", err=None, lock_seconds=None)

    username = (request.form.get("username") or "").strip()
    password = request.form.get("password") or ""
    remember_me = truthy(request.form.get("remember_me", "0"))
    throttle_keys = login_throttle_keys(username)
    locked, seconds_left = login_throttle_status(throttle_keys)
    if locked:
        return render_template("login.html", err="Too many attempts.", lock_seconds=seconds_left)

    user = User.query.filter_by(username=username).first()
    if not user or not check_password_hash(user.password_hash, password):
        login_throttle_failed(throttle_keys)
        return render_template("login.html", err="Invalid username or password.", lock_seconds=None)
    if user.role == "disabled":
        login_throttle_failed(throttle_keys)
        return render_template("login.html", err="Account is deactivated. Please contact your coach.", lock_seconds=None)
    login_throttle_success(throttle_keys)

    now = utc_now()
    user.last_login_at = now
    user.last_seen_at = now
    db.session.commit()

    session["user_id"] = user.id
    session["role"] = user.role
    session["client_id"] = user.client_id
    session.permanent = remember_me

    if user.role == "admin":
        return redirect(url_for("index"))
    if user.must_change_password:
        return redirect(url_for(
            "client_profile",
            client_id=user.client_id,
            tab="info",
            err="You must set a new password before using other tabs."
        ))
    return redirect(url_for("client_profile", client_id=user.client_id, tab="info"))


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# =========================
# Main Routes
# =========================
@app.route("/")
@login_required
def index():
    if not is_admin():
        # client users go straight to their profile
        cid = current_client_id()
        return redirect(url_for("client_profile", client_id=cid, tab="info"))

    show_inactive = truthy(request.args.get("show_inactive", "0"))
    clients_query = Client.query.order_by(Client.created_at.desc())
    clients = clients_query.all() if show_inactive else clients_query.filter_by(is_active=True).all()
    hidden_clients_count = Client.query.filter_by(is_active=False).count()
    payment_alerts = []
    for c in clients:
        if not c.is_active:
            continue
        _spw, status = get_current_plan(c.id)
        if not status:
            continue
        days_left = status["days_left"]
        if days_left < 0:
            payment_alerts.append({
                "client": c,
                "tone": "overdue",
                "label": f"Overdue by {abs(days_left)} day(s)",
                "days_left": days_left,
            })
        elif days_left <= 7:
            payment_alerts.append({
                "client": c,
                "tone": "due-soon",
                "label": f"Due in {days_left} day(s)",
                "days_left": days_left,
            })
    payment_alerts.sort(key=lambda x: x["days_left"])
    err = request.args.get("err")
    msg = request.args.get("msg")
    return render_template(
        "index.html",
        clients=clients,
        err=err,
        msg=msg,
        payment_alerts=payment_alerts,
        show_inactive=show_inactive,
        hidden_clients_count=hidden_clients_count,
    )


@app.route("/admin/budget")
@login_required
def budget_tracker():
    if not is_admin():
        return "Forbidden", 403

    requested = parse_iso_date(request.args.get("week"))
    ws = week_start(requested or date.today())
    view = build_budget_week_view(ws)

    return render_template(
        "budget.html",
        title="Weekly Budget",
        view=view,
        prev_week=(ws - timedelta(days=7)).isoformat(),
        next_week=(ws + timedelta(days=7)).isoformat(),
        this_week=week_start(date.today()).isoformat(),
        is_current_week=(ws == week_start(date.today())),
        all_time_savings=compute_all_time_savings(),
        err=request.args.get("err"),
        msg=request.args.get("msg"),
    )


@app.route("/admin/budget/set-weekly", methods=["POST"])
@login_required
def set_weekly_budget():
    if not is_admin():
        return "Forbidden", 403

    ws = week_start(parse_iso_date(request.form.get("week_start")) or date.today())
    raw = (request.form.get("weekly_budget") or "").strip()
    try:
        amount = round(max(0.0, float(raw)))
    except ValueError:
        return redirect(url_for("budget_tracker", week=ws.isoformat(), err="Enter a valid weekly budget."))

    week = get_budget_week(ws)
    if not week:
        db.session.add(BudgetWeek(week_start=ws, weekly_budget=amount))
    else:
        week.weekly_budget = amount
    db.session.commit()
    return redirect(url_for("budget_tracker", week=ws.isoformat(), msg="Weekly budget updated."))


@app.route("/admin/budget/log-day", methods=["POST"])
@login_required
def log_budget_day():
    if not is_admin():
        return "Forbidden", 403

    is_ajax = request.headers.get("X-Budget-Ajax") == "1"

    d = parse_iso_date(request.form.get("day"))
    if not d:
        return redirect(url_for("budget_tracker"))

    raw = (request.form.get("spent") or "").strip()
    try:
        amount = round(max(0.0, float(raw)))
    except ValueError:
        if is_ajax:
            return jsonify({"ok": False, "error": "Enter a valid amount spent."}), 400
        return redirect(url_for("budget_tracker", week=week_start(d).isoformat(), err="Enter a valid amount spent."))

    entry = BudgetDay.query.filter_by(day=d).first()
    if not entry:
        db.session.add(BudgetDay(day=d, spent=amount))
    else:
        entry.spent = amount
    db.session.commit()

    if is_ajax:
        return jsonify(budget_day_ajax_payload(d))
    return redirect(url_for(
        "budget_tracker", week=week_start(d).isoformat(),
        msg=f"Logged {amount} MKD spent on {d.strftime('%b %d')}.",
    ))


@app.route("/admin/budget/clear-day", methods=["POST"])
@login_required
def clear_budget_day():
    if not is_admin():
        return "Forbidden", 403

    is_ajax = request.headers.get("X-Budget-Ajax") == "1"

    d = parse_iso_date(request.form.get("day"))
    if d:
        entry = BudgetDay.query.filter_by(day=d).first()
        if entry:
            db.session.delete(entry)
            db.session.commit()

    if is_ajax and d:
        return jsonify(budget_day_ajax_payload(d))
    return redirect(url_for("budget_tracker", week=week_start(d).isoformat() if d else None, msg="Entry cleared."))


@app.route("/client/<int:client_id>")
@login_required
def client_profile(client_id):
    # access control
    if not is_admin():
        if current_client_id() != client_id:
            return "Forbidden", 403

    client = get_or_404(Client, client_id)
    seed_reference_foods()
    tab = request.args.get("tab", "info")
    nutrition_date = parse_iso_date(request.args.get("nutrition_date")) or date.today()
    strength_exercise = (request.args.get("strength_exercise") or "").strip()
    poster_start_raw = None
    poster_end_raw = None
    poster_lifts = request.args.getlist("poster_lifts")
    err = request.args.get("err")
    msg = request.args.get("msg")
    viewer = current_user()
    # The old standalone "sessions" tab was folded into "info".
    if tab == "sessions":
        return redirect(url_for("client_profile", client_id=client.id, tab="info"))
    # Keep client navigation on allowed tabs only.
    if not is_admin() and tab in ("payments", "strength"):
        return redirect(url_for("client_profile", client_id=client.id, tab="info"))
    if not is_admin() and viewer and viewer.must_change_password and tab != "info":
        return redirect(url_for(
            "client_profile",
            client_id=client.id,
            tab="info",
            err="You must set a new password before using other tabs."
        ))

    # Stats
    all_measurements = (
        Measurement.query.filter_by(client_id=client.id)
        .order_by(Measurement.date.desc())
        .all()
    )

    def has_body_measurements(m: Measurement) -> bool:
        return any(
            v is not None
            for v in (
                m.chest, m.waist, m.stomach, m.glutes,
                m.arm_left, m.arm_right, m.quad_left, m.quad_right,
                m.calf_left, m.calf_right,
            )
        )

    measurements = [m for m in all_measurements if has_body_measurements(m)]
    latest = measurements[0] if measurements else None

    # Weight-only/weight history data
    weight_points_asc = (
        Measurement.query.filter_by(client_id=client.id)
        .filter(Measurement.weight.isnot(None))
        .order_by(Measurement.date.asc())
        .all()
    )
    weight_latest = weight_points_asc[-1] if weight_points_asc else None
    weight_measurements = list(reversed(weight_points_asc))
    weight_change = None
    if len(weight_points_asc) >= 2:
        weight_change = round(weight_points_asc[-1].weight - weight_points_asc[0].weight, 1)

    # The weight graph builds one SVG point per entry client-side; for a client
    # logging very frequently that's an unbounded string-build + DOM write that
    # visibly stutters on the one tab that renders it. A recent-trend window is
    # both fast and more readable than plotting years of daily points at once.
    CHART_MAX_POINTS = 200
    weight_chart_points = "".join(
        f"{m.date.strftime('%Y-%m-%d')},{m.weight};" for m in weight_points_asc[-CHART_MAX_POINTS:]
    )

    # Clients who log frequently can accumulate hundreds of entries; rendering a
    # full edit modal per row (see stats.html) made the page balloon in DOM size.
    # Cap what's actually listed/editable per view and let "Show more" raise it.
    HISTORY_PAGE_SIZE = 50

    def parse_history_limit(param_name):
        try:
            value = int(request.args.get(param_name, HISTORY_PAGE_SIZE))
        except (TypeError, ValueError):
            value = HISTORY_PAGE_SIZE
        return max(HISTORY_PAGE_SIZE, value)

    weight_history_limit = parse_history_limit("weight_limit")
    measurement_history_limit = parse_history_limit("measurement_limit")
    weight_history_display = weight_measurements[:weight_history_limit]
    measurements_display = measurements[:measurement_history_limit]
    weight_history_has_more = len(weight_measurements) > weight_history_limit
    measurements_history_has_more = len(measurements) > measurement_history_limit

    # Sessions / Payments Plan
    sessions_per_week_from_payments, current_status = get_current_plan(client.id)
    sessions_per_week = (
        sessions_per_week_from_payments
        if sessions_per_week_from_payments is not None
        else (client.weekly_sessions or 0)
    )

    used_this_week, remaining, bonus, allowed = compute_sessions(client, sessions_per_week)

    payment_status_label = None
    payment_status_tone = None
    if current_status:
        days_left = current_status["days_left"]
        if days_left < 0:
            payment_status_label = "Overdue"
            payment_status_tone = "overdue"
        elif days_left <= 7:
            payment_status_label = "Due soon"
            payment_status_tone = "due-soon"
        else:
            payment_status_label = "Active"
            payment_status_tone = "active"

    sessions = (
        SessionLog.query.filter_by(client_id=client.id)
        .order_by(SessionLog.date.desc())
        .limit(50)
        .all()
    )
    session_calendar = build_session_calendar(sessions)
    total_sessions = SessionLog.query.filter_by(client_id=client.id).count()
    thirty_days_ago_dt = utc_now() - timedelta(days=30)
    sessions_30d = (
        SessionLog.query.filter_by(client_id=client.id)
        .filter(SessionLog.date >= thirty_days_ago_dt)
        .count()
    )
    target_30d = max((sessions_per_week or 0) * 4, 1)
    adherence_30d = min(int((sessions_30d / target_30d) * 100), 100)
    weight_30d_points = (
        Measurement.query.filter_by(client_id=client.id)
        .filter(Measurement.weight.isnot(None))
        .filter(Measurement.date >= thirty_days_ago_dt)
        .order_by(Measurement.date.asc())
        .all()
    )
    weight_change_30d = None
    if len(weight_30d_points) >= 2:
        weight_change_30d = round(weight_30d_points[-1].weight - weight_30d_points[0].weight, 1)

    # Payments view
    payments = (
        Payment.query.filter_by(client_id=client.id)
        .order_by(Payment.start_date.desc(), Payment.paid_on.desc(), Payment.id.desc())
        .all()
    )

    today = date.today()
    payments_view = []
    for idx, p in enumerate(payments):
        due = payment_due_date(p)
        days_left = (due - today).days
        payments_view.append({
            "p": p,
            "due": due,
            "days_left": days_left,
            "is_current": idx == 0,
        })

    client_user = (
        User.query
        .filter_by(client_id=client.id)
        .filter(User.role != "admin")
        .order_by(User.id.desc())
        .first()
    )
    client_online_now = False
    client_last_seen_display = "-"
    client_last_login_display = "-"
    if client_user:
        client_last_seen_display = humanize_last_seen(client_user.last_seen_at)
        if client_user.last_login_at:
            client_last_login_display = client_user.last_login_at.strftime("%d/%m/%Y %H:%M UTC")
        if client_user.last_seen_at and client_user.role != "disabled":
            seen_at = client_user.last_seen_at
            if seen_at.tzinfo is not None:
                seen_at = seen_at.astimezone(timezone.utc).replace(tzinfo=None)
            client_online_now = (utc_now() - seen_at).total_seconds() <= 300
    appointments = (
        Appointment.query.filter_by(client_id=client.id)
        .order_by(Appointment.scheduled_for.asc(), Appointment.id.asc())
        .all()
    )
    pending_today_appointments = []
    today = date.today()
    if not is_admin():
        pending_today_appointments = [
            a for a in appointments
            if a.status == "requested" and a.scheduled_for.date() == today
        ]
    photos = (
        ProgressPhoto.query.filter_by(client_id=client.id)
        .order_by(ProgressPhoto.created_at.desc(), ProgressPhoto.id.desc())
        .all()
    )
    goals = (
        ClientGoal.query.filter_by(client_id=client.id)
        .order_by(ClientGoal.created_at.desc(), ClientGoal.id.desc())
        .all()
    )
    foods = (
        FoodItem.query
        .filter(or_(FoodItem.client_id.is_(None), FoodItem.client_id == client.id))
        .order_by(FoodItem.name.asc(), FoodItem.brand.asc(), FoodItem.id.asc())
        .all()
    )
    food_usage_rows = (
        db.session.query(
            FoodLogEntry.food_id,
            func.count(FoodLogEntry.id),
            func.max(FoodLogEntry.created_at),
        )
        .filter(FoodLogEntry.client_id == client.id)
        .group_by(FoodLogEntry.food_id)
        .all()
    )
    food_usage_stats = {
        food_id: {
            "count": int(count or 0),
            "last_used": last_used.isoformat() if last_used else "",
        }
        for food_id, count, last_used in food_usage_rows
    }
    food_logs = (
        FoodLogEntry.query.filter_by(client_id=client.id, logged_for=nutrition_date)
        .order_by(FoodLogEntry.created_at.desc(), FoodLogEntry.id.desc())
        .all()
    )
    nutrition_summary = build_nutrition_summary(food_logs, client.daily_calorie_target)
    nutrition_search_results = get_nutrition_search_results(client.id)
    nutrition_label_draft = get_nutrition_label_draft(client.id)
    strength_entries = (
        StrengthLogEntry.query.filter_by(client_id=client.id)
        .order_by(StrengthLogEntry.logged_at.asc(), StrengthLogEntry.id.asc())
        .all()
    )
    strength_summary = build_strength_summary(strength_entries, strength_exercise)
    poster_start, poster_end = parse_strength_date_range(poster_start_raw, poster_end_raw, strength_entries)
    strength_poster = build_strength_poster_data(
        strength_entries,
        weight_points_asc,
        poster_start,
        poster_end,
        poster_lifts,
    )
    if is_admin():
        notes = (
            ClientNote.query.filter_by(client_id=client.id)
            .order_by(ClientNote.created_at.desc(), ClientNote.id.desc())
            .all()
        )
    else:
        notes = (
            ClientNote.query.filter_by(client_id=client.id, is_private=False)
            .order_by(ClientNote.created_at.desc(), ClientNote.id.desc())
            .all()
        )

    must_change_password = bool(client_user.must_change_password) if client_user else False
    milestones = []
    if total_sessions >= 10:
        milestones.append("10+ sessions completed")
    if total_sessions >= 25:
        milestones.append("25+ sessions consistency")
    if weight_change is not None and weight_change <= -2:
        milestones.append(f"Weight down {abs(weight_change):.1f} kg")
    if not milestones:
        milestones.append("First milestone pending")

    return render_template(
        "client.html",
        client=client,
        tab=tab,
        page_class="nutrition-shell" if tab == "nutrition" else ("strength-shell" if tab == "strength" else ""),
        err=err,
        msg=msg,
        is_admin=is_admin(),

        # stats
        measurements=measurements,
        measurements_display=measurements_display,
        measurement_history_limit=measurement_history_limit,
        measurements_history_has_more=measurements_history_has_more,
        latest=latest,
        weight_latest=weight_latest,
        weight_measurements=weight_measurements,
        weight_history_display=weight_history_display,
        weight_history_limit=weight_history_limit,
        weight_history_has_more=weight_history_has_more,
        weight_change=weight_change,
        weight_chart_points=weight_chart_points,

        # sessions
        sessions=sessions,
        session_calendar=session_calendar,
        total_sessions=total_sessions,
        sessions_30d=sessions_30d,
        adherence_30d=adherence_30d,
        weight_change_30d=weight_change_30d,
        used_this_week=used_this_week,
        remaining=remaining,
        bonus=bonus,
        allowed=allowed,
        sessions_per_week=sessions_per_week,

        # payments
        payments_view=payments_view,
        current_status=current_status,
        payment_status_label=payment_status_label,
        payment_status_tone=payment_status_tone,
        notes=notes,
        appointments=appointments,
        pending_today_appointments=pending_today_appointments,
        photos=photos,
        goals=goals,
        foods=foods,
        food_usage_stats=food_usage_stats,
        food_logs=food_logs,
        nutrition_date=nutrition_date,
        nutrition_today=date.today(),
        nutrition_day_title=nutrition_day_label(nutrition_date),
        nutrition_summary=nutrition_summary,
        nutrition_search_results=nutrition_search_results,
        nutrition_label_draft=nutrition_label_draft,
        label_scan_enabled=label_scan_enabled(),
        label_scan_provider=label_scan_provider_label(),
        strength_summary=strength_summary,
        strength_poster=strength_poster,
        strength_poster_selected_lifts=poster_lifts,
        estimate_one_rep_max=estimate_one_rep_max,
        client_user=client_user,
        client_online_now=client_online_now,
        client_last_seen_display=client_last_seen_display,
        client_last_login_display=client_last_login_display,
        must_change_password=must_change_password,
        milestones=milestones,
    )


@app.route("/export/clients.csv")
@login_required
def export_clients_csv():
    if not is_admin():
        return "Forbidden", 403

    clients = Client.query.order_by(Client.created_at.desc()).all()
    log_security_event("export_clients_csv", f"rows={len(clients)}")
    rows = [
        [
            c.id,
            c.name,
            c.phone or "",
            c.plan or "",
            c.weekly_sessions or 0,
            c.created_at.strftime("%Y-%m-%d %H:%M:%S") if c.created_at else "",
        ]
        for c in clients
    ]
    return csv_download_response(
        "clients_export.csv",
        ["id", "name", "phone", "plan", "weekly_sessions", "created_at"],
        rows,
    )


@app.route("/export/sessions.csv")
@login_required
def export_sessions_csv():
    if not is_admin():
        return "Forbidden", 403

    clients = {c.id: c.name for c in Client.query.all()}
    sessions = SessionLog.query.order_by(SessionLog.date.desc()).all()
    log_security_event("export_sessions_csv", f"rows={len(sessions)}")
    rows = [
        [
            s.id,
            s.client_id,
            clients.get(s.client_id, ""),
            s.date.strftime("%Y-%m-%d %H:%M:%S") if s.date else "",
            s.note or "",
        ]
        for s in sessions
    ]
    return csv_download_response(
        "sessions_export.csv",
        ["id", "client_id", "client_name", "date", "note"],
        rows,
    )


@app.route("/export/payments.csv")
@login_required
def export_payments_csv():
    if not is_admin():
        return "Forbidden", 403

    clients = {c.id: c.name for c in Client.query.all()}
    payments = Payment.query.order_by(Payment.start_date.desc()).all()
    log_security_event("export_payments_csv", f"rows={len(payments)}")
    rows = []
    for p in payments:
        due = payment_due_date(p)
        rows.append([
            p.id,
            p.client_id,
            clients.get(p.client_id, ""),
            p.start_date.strftime("%Y-%m-%d") if p.start_date else "",
            p.months,
            p.sessions_per_week,
            p.monthly_price,
            p.amount_paid,
            due.strftime("%Y-%m-%d"),
            p.note or "",
            p.paid_on.strftime("%Y-%m-%d %H:%M:%S") if p.paid_on else "",
        ])

    return csv_download_response(
        "payments_export.csv",
        [
            "id",
            "client_id",
            "client_name",
            "start_date",
            "months",
            "sessions_per_week",
            "monthly_price",
            "amount_paid",
            "due_date",
            "note",
            "paid_on",
        ],
        rows,
    )


@app.route("/backup/database")
@login_required
def backup_database():
    if not is_admin():
        return "Forbidden", 403

    db_file_path = resolve_db_file_path()
    if not db_file_path:
        return "Database file not found.", 404

    log_security_event("backup_database", f"path={db_file_path}")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return send_file(
        db_file_path,
        as_attachment=True,
        download_name=f"trainer_backup_{timestamp}.db",
        mimetype="application/octet-stream",
    )


@app.route("/admin/import-capnutra", methods=["POST"])
@login_required
def import_capnutra_admin():
    if not is_admin():
        return "Forbidden", 403

    try:
        stats = import_capnutra_foods(fetch_capnutra_food_library(), prune_missing=True)
    except Exception:
        app.logger.exception("CAPNUTRA import failed")
        return redirect(url_for(
            "index",
            err="CAPNUTRA import failed. Try again in a minute, or check the deployment logs.",
        ))

    return redirect(url_for(
        "index",
        msg=(
            "CAPNUTRA import complete: "
            f"{stats['created']} created, {stats['updated']} updated, "
            f"{stats['removed']} removed, {stats['total']} active."
        ),
    ))


@app.route("/client/<int:client_id>/report.pdf")
@login_required
def export_client_report_pdf(client_id):
    if not is_admin() and current_client_id() != client_id:
        return "Forbidden", 403

    client = get_or_404(Client, client_id)
    sessions_per_week_from_payments, current_status = get_current_plan(client.id)
    sessions_per_week = (
        sessions_per_week_from_payments
        if sessions_per_week_from_payments is not None
        else (client.weekly_sessions or 0)
    )
    used_this_week, remaining, bonus, allowed = compute_sessions(client, sessions_per_week)

    latest_weight = (
        Measurement.query.filter_by(client_id=client.id)
        .filter(Measurement.weight.isnot(None))
        .order_by(Measurement.date.desc())
        .first()
    )
    total_sessions = SessionLog.query.filter_by(client_id=client.id).count()

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
    except Exception:
        return "PDF export requires reportlab package.", 500

    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    y = height - 40

    def line(text, dy=20):
        nonlocal y
        c.drawString(40, y, text)
        y -= dy
        if y < 60:
            c.showPage()
            y = height - 40

    c.setFont("Helvetica-Bold", 15)
    line(f"Client Report - {client.name}", 26)
    c.setFont("Helvetica", 11)
    line(f"Phone: {client.phone or '-'}")
    line(f"Plan: {client.plan or '-'}")
    line(f"Sessions/week: {sessions_per_week} | Used this week: {used_this_week}/{allowed} | Remaining: {remaining}")
    line(f"Total sessions logged: {total_sessions}")
    if latest_weight:
        line(f"Latest weight: {latest_weight.weight} kg ({latest_weight.date.strftime('%d/%m/%Y')})")
    else:
        line("Latest weight: -")

    if current_status:
        line(
            f"Payment status: paid {current_status['amount_paid']} MKD | start {current_status['start_date'].strftime('%d/%m/%Y')} | due {current_status['paid_until'].strftime('%d/%m/%Y')} | days left {current_status['days_left']}"
        )
    else:
        line("Payment status: no active payment")

    notes = (
        ClientNote.query.filter_by(client_id=client.id)
        .order_by(ClientNote.created_at.desc())
        .limit(8)
        .all()
    )
    line("Recent notes:", 24)
    for n in notes:
        privacy = "Private" if n.is_private else "Shared"
        line(f"- [{privacy}] {n.created_at.strftime('%d/%m/%Y %H:%M')} - {n.text}", 16)

    c.save()
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"client_report_{client.id}.pdf",
        mimetype="application/pdf",
    )


# =========================
# Admin: Add/Delete Client
# =========================
@app.route("/admin/change-credentials", methods=["POST"])
@login_required
def change_admin_credentials():
    if not is_admin():
        return "Forbidden", 403

    admin_user = User.query.filter_by(id=session.get("user_id"), role="admin").first()
    if not admin_user:
        return redirect(url_for("index", err="Admin user not found."))

    current_password = request.form.get("current_password") or ""
    new_username = (request.form.get("new_username") or "").strip()
    new_password = request.form.get("new_password") or ""
    confirm_password = request.form.get("confirm_password") or ""

    if not check_password_hash(admin_user.password_hash, current_password):
        return redirect(url_for("index", err="Current password is incorrect."))

    if not new_username:
        return redirect(url_for("index", err="New username is required."))
    if len(new_username) > 80:
        return redirect(url_for("index", err="Username is too long."))
    if not re.fullmatch(r"[A-Za-z0-9._-]+", new_username):
        return redirect(url_for("index", err="Username can only contain letters, numbers, dot, underscore, and hyphen."))

    existing_username = User.query.filter_by(username=new_username).first()
    if existing_username and existing_username.id != admin_user.id:
        return redirect(url_for("index", err="Username is already taken."))

    changing_password = bool(new_password or confirm_password)
    if changing_password:
        if len(new_password) < 6:
            return redirect(url_for("index", err="New password must be at least 6 characters."))
        if new_password != confirm_password:
            return redirect(url_for("index", err="New password and confirmation do not match."))
        admin_user.password_hash = generate_password_hash(new_password)

    admin_user.username = new_username
    db.session.commit()

    return redirect(url_for("index", msg="Admin credentials updated."))


@app.route("/add", methods=["POST"])
@login_required
def add_client():
    if not is_admin():
        return "Forbidden", 403

    name = request.form["name"].strip()
    phone = parse_phone(request.form.get("phone", ""))
    if phone is None:
        return redirect(url_for("index", err="Phone can contain only digits and an optional leading +."))
    plan = request.form.get("plan", "").strip()

    new_client = Client(name=name, phone=phone, plan=plan, weekly_sessions=0)
    db.session.add(new_client)
    db.session.commit()
    return redirect(url_for("index"))


@app.route("/delete/<int:client_id>", methods=["POST"])
@login_required
def delete_client(client_id):
    if not is_admin():
        return "Forbidden", 403

    client = get_or_404(Client, client_id)
    db.session.delete(client)
    db.session.commit()
    return redirect(url_for("index"))


@app.route("/client/<int:client_id>/status", methods=["POST"])
@login_required
def update_client_status(client_id):
    if not is_admin():
        return "Forbidden", 403

    client = get_or_404(Client, client_id)
    action = (request.form.get("action") or "").strip().lower()
    if action == "hide":
        client.is_active = False
    elif action == "show":
        client.is_active = True
    else:
        return redirect(url_for("index", err="Invalid client status action."))
    db.session.commit()
    show_inactive = "1" if truthy(request.form.get("show_inactive", "0")) else "0"
    return redirect(url_for("index", show_inactive=show_inactive, msg="Client status updated."))


# =========================
# Client updates
# =========================
@app.route("/client/<int:client_id>/update-admin", methods=["POST"])
@login_required
def update_client_admin(client_id):
    if not is_admin():
        return "Forbidden", 403

    client = get_or_404(Client, client_id)
    client.name = request.form.get("name", client.name).strip()
    phone = parse_phone(request.form.get("phone", client.phone or ""))
    if phone is None:
        return redirect(url_for("client_profile", client_id=client.id, tab="info", err="Phone can contain only digits and an optional leading +."))
    client.phone = phone
    client.plan = request.form.get("plan", client.plan or "").strip()
    client.weekly_sessions = to_int(request.form.get("weekly"), default=0)

    db.session.commit()
    return redirect(url_for("client_profile", client_id=client.id, tab="info", msg="Saved"))


@app.route("/client/<int:client_id>/update-phone", methods=["POST"])
@login_required
def update_client_phone(client_id):
    # client can only update own phone, admin can update any
    if not is_admin() and current_client_id() != client_id:
        return "Forbidden", 403

    client = get_or_404(Client, client_id)
    phone = parse_phone(request.form.get("phone"))
    if phone is None:
        return redirect(url_for("client_profile", client_id=client.id, tab="info", err="Phone can contain only digits and an optional leading +."))
    client.phone = phone
    db.session.commit()
    return redirect(url_for("client_profile", client_id=client.id, tab="info", msg="Phone updated"))


@app.route("/client/<int:client_id>/update", methods=["POST"], endpoint="update_client")
@login_required
def update_client_alias(client_id):
    # Backwards compatible endpoint name.
    # Admin can update everything, clients only phone.
    if is_admin():
        return update_client_admin(client_id)
    return update_client_phone(client_id)


@app.route("/client/<int:client_id>/notes/add", methods=["POST"])
@login_required
def add_client_note(client_id):
    if not is_admin() and current_client_id() != client_id:
        return "Forbidden", 403

    client = get_or_404(Client, client_id)
    text = (request.form.get("text") or "").strip()
    is_private = truthy(request.form.get("is_private", "0"))
    if not text:
        return redirect(url_for("client_profile", client_id=client.id, tab="info", err="Note cannot be empty."))
    if len(text) > 500:
        return redirect(url_for("client_profile", client_id=client.id, tab="info", err="Note is too long (max 500 chars)."))
    if not is_admin():
        is_private = False

    note = ClientNote(
        client_id=client.id,
        text=text,
        is_private=is_private,
        created_by_role="admin" if is_admin() else "client",
    )
    db.session.add(note)
    db.session.commit()
    return redirect(url_for("client_profile", client_id=client.id, tab="info", msg="Note added."))


@app.route("/client/<int:client_id>/notes/delete/<int:note_id>", methods=["POST"])
@login_required
def delete_client_note(client_id, note_id):
    if not is_admin():
        return "Forbidden", 403
    note = get_or_404(ClientNote, note_id)
    if note.client_id != client_id:
        abort(404)
    db.session.delete(note)
    db.session.commit()
    return redirect(url_for("client_profile", client_id=client_id, tab="info", msg="Note deleted."))


@app.route("/client/<int:client_id>/goals/add", methods=["POST"])
@login_required
def add_client_goal(client_id):
    if not is_admin() and current_client_id() != client_id:
        return "Forbidden", 403
    client = get_or_404(Client, client_id)
    title = (request.form.get("title") or "").strip()
    if not title:
        return redirect(url_for("client_profile", client_id=client.id, tab="info", err="Goal title is required."))
    g = ClientGoal(
        client_id=client.id,
        title=title,
        goal_type="custom",
        unit=None,
        target_value=None,
        current_value=None,
        target_date=None,
        status="active",
        note="",
    )
    db.session.add(g)
    db.session.commit()
    return redirect(url_for("client_profile", client_id=client.id, tab="info", msg="Goal added."))


@app.route("/client/<int:client_id>/goals/update/<int:goal_id>", methods=["POST"])
@login_required
def update_client_goal(client_id, goal_id):
    if not is_admin():
        return "Forbidden", 403
    g = get_or_404(ClientGoal, goal_id)
    if g.client_id != client_id:
        abort(404)
    g.current_value = to_float(request.form.get("current_value"))
    g.goal_type = normalize_goal_type(request.form.get("goal_type") or g.goal_type)
    if request.form.get("unit") is not None:
        g.unit = normalize_goal_unit(request.form.get("unit")) or default_goal_unit(g.goal_type)
    status = (request.form.get("status") or "active").strip().lower()
    if status not in ("active", "completed", "paused"):
        status = "active"
    g.status = status
    if "note" in request.form:
        g.note = (request.form.get("note") or "").strip()
    db.session.commit()
    return redirect(url_for("client_profile", client_id=client_id, tab="info", msg="Goal updated."))


@app.route("/client/<int:client_id>/goals/delete/<int:goal_id>", methods=["POST"])
@login_required
def delete_client_goal(client_id, goal_id):
    if not is_admin():
        return "Forbidden", 403
    g = get_or_404(ClientGoal, goal_id)
    if g.client_id != client_id:
        abort(404)
    db.session.delete(g)
    db.session.commit()
    return redirect(url_for("client_profile", client_id=client_id, tab="info", msg="Goal deleted."))


@app.route("/client/<int:client_id>/nutrition/target", methods=["POST"])
@login_required
def update_calorie_target(client_id):
    if not is_admin() and current_client_id() != client_id:
        return "Forbidden", 403

    client = get_or_404(Client, client_id)
    target_raw = (request.form.get("daily_calorie_target") or "").strip()
    nutrition_date = parse_iso_date(request.form.get("nutrition_date")) or date.today()
    if not target_raw:
        client.daily_calorie_target = None
        db.session.commit()
        return nutrition_redirect(client.id, msg="Daily calorie target cleared.", logged_for=nutrition_date)

    try:
        target = int(target_raw)
    except ValueError:
        return nutrition_redirect(client.id, err="Daily calorie target must be a whole number.", logged_for=nutrition_date)
    if target <= 0:
        return nutrition_redirect(client.id, err="Daily calorie target must be greater than 0.", logged_for=nutrition_date)

    client.daily_calorie_target = target
    db.session.commit()
    return nutrition_redirect(client.id, msg="Daily calorie target updated.", logged_for=nutrition_date)


@app.route("/client/<int:client_id>/nutrition/update-target-ajax", methods=["POST"])
@login_required
def update_calorie_target_ajax(client_id):
    if not is_admin() and current_client_id() != client_id:
        return jsonify({"success": False, "error": "Forbidden"}), 403

    client = get_or_404(Client, client_id)
    target_raw = (request.form.get("daily_calorie_target") or "").strip()
    nutrition_date = parse_iso_date(request.form.get("nutrition_date")) or date.today()

    if not target_raw:
        client.daily_calorie_target = None
        db.session.commit()
        # Get updated nutrition summary
        food_logs = (
            FoodLogEntry.query.filter_by(client_id=client.id, logged_for=nutrition_date)
            .order_by(FoodLogEntry.created_at.desc(), FoodLogEntry.id.desc())
            .all()
        )
        nutrition_summary = build_nutrition_summary(food_logs, client.daily_calorie_target)
        return jsonify({
            "success": True,
            "message": "Daily calorie target cleared.",
            "nutrition_summary": nutrition_summary_payload(nutrition_summary),
        })

    try:
        target = int(target_raw)
    except ValueError:
        return jsonify({"success": False, "error": "Daily calorie target must be a whole number."}), 400
    if target <= 0:
        return jsonify({"success": False, "error": "Daily calorie target must be greater than 0."}), 400

    client.daily_calorie_target = target
    db.session.commit()

    # Get updated nutrition summary
    food_logs = (
        FoodLogEntry.query.filter_by(client_id=client.id, logged_for=nutrition_date)
        .order_by(FoodLogEntry.created_at.desc(), FoodLogEntry.id.desc())
        .all()
    )
    nutrition_summary = build_nutrition_summary(food_logs, client.daily_calorie_target)

    return jsonify({
        "success": True,
        "message": "Daily calorie target updated!",
        "nutrition_summary": nutrition_summary_payload(nutrition_summary),
    })


@app.route("/client/<int:client_id>/nutrition/day-summary-ajax")
@login_required
def nutrition_day_summary_ajax(client_id):
    if not is_admin() and current_client_id() != client_id:
        return jsonify({"success": False, "error": "Forbidden"}), 403

    client = get_or_404(Client, client_id)
    nutrition_date = parse_iso_date(request.args.get("nutrition_date")) or date.today()
    if nutrition_date > date.today():
        return jsonify({"success": False, "error": "Future days are not available."}), 400

    food_logs = (
        FoodLogEntry.query.filter_by(client_id=client.id, logged_for=nutrition_date)
        .order_by(FoodLogEntry.created_at.desc(), FoodLogEntry.id.desc())
        .all()
    )
    nutrition_summary = build_nutrition_summary(food_logs, client.daily_calorie_target)
    return jsonify({
        "success": True,
        "nutrition_date": nutrition_date.isoformat(),
        "date_label": nutrition_day_label(nutrition_date),
        "date_display": nutrition_date.strftime("%d/%m/%Y"),
        "is_today": nutrition_date == date.today(),
        "nutrition_summary": nutrition_summary_payload(nutrition_summary),
    })


@app.route("/client/<int:client_id>/nutrition/foods/add", methods=["POST"])
@login_required
def add_food_item(client_id):
    if not is_admin() and current_client_id() != client_id:
        return "Forbidden", 403

    client = get_or_404(Client, client_id)
    nutrition_date = parse_iso_date(request.form.get("nutrition_date")) or date.today()
    name = (request.form.get("name") or "").strip()
    if not name:
        return nutrition_redirect(client.id, err="Food name is required.", logged_for=nutrition_date)

    calories_per_100g, err = parse_positive_float(request.form.get("calories_per_100g"), "Calories per 100g")
    if err:
        return nutrition_redirect(client.id, err=err, logged_for=nutrition_date)

    protein_per_100g, err = parse_positive_float(request.form.get("protein_per_100g"), "Protein per 100g", allow_zero=True)
    if err:
        return nutrition_redirect(client.id, err=err, logged_for=nutrition_date)
    carbs_per_100g, err = parse_positive_float(request.form.get("carbs_per_100g"), "Carbs per 100g", allow_zero=True)
    if err:
        return nutrition_redirect(client.id, err=err, logged_for=nutrition_date)
    fat_per_100g, err = parse_positive_float(request.form.get("fat_per_100g"), "Fat per 100g", allow_zero=True)
    if err:
        return nutrition_redirect(client.id, err=err, logged_for=nutrition_date)

    food = FoodItem(
        client_id=client.id,
        name=name[:120],
        brand=((request.form.get("brand") or "").strip() or None),
        serving_label=((request.form.get("serving_label") or "").strip() or None),
        source="manual",
        calories_per_100g=calories_per_100g,
        protein_per_100g=protein_per_100g,
        carbs_per_100g=carbs_per_100g,
        fat_per_100g=fat_per_100g,
    )
    db.session.add(food)
    db.session.commit()
    set_nutrition_label_draft(client.id, None)
    if request.form.get("save_action") == "save_and_log":
        quantity_grams, err = parse_positive_float(request.form.get("log_quantity_grams"), "Quantity in grams")
        if err:
            return nutrition_redirect(client.id, msg="Food saved to your library. Add a quantity to log it.", logged_for=nutrition_date, anchor="nutritionLogForm")

        factor = quantity_grams / 100
        food_label = food.name if not food.brand else f"{food.name} ({food.brand})"
        log_entry = FoodLogEntry(
            client_id=client.id,
            food_id=food.id,
            logged_for=nutrition_date,
            meal_type=normalize_meal_type(request.form.get("log_meal_type")),
            quantity_grams=quantity_grams,
            food_name=food_label[:160],
            calories=round(food.calories_per_100g * factor, 1),
            protein=round(food.protein_per_100g * factor, 1),
            carbs=round(food.carbs_per_100g * factor, 1),
            fat=round(food.fat_per_100g * factor, 1),
            note=((request.form.get("log_note") or "").strip() or None),
        )
        db.session.add(log_entry)
        db.session.commit()
        return nutrition_redirect(client.id, msg="Food saved and logged.", logged_for=nutrition_date, anchor="nutritionLogForm")

    return nutrition_redirect(client.id, msg="Food saved to your library.", logged_for=nutrition_date)


@app.route("/client/<int:client_id>/nutrition/search", methods=["POST"])
@login_required
def search_food_sources(client_id):
    if not is_admin() and current_client_id() != client_id:
        return "Forbidden", 403

    client = get_or_404(Client, client_id)
    seed_reference_foods()
    logged_for = parse_iso_date(request.form.get("nutrition_date")) or date.today()
    source = (request.form.get("source") or "").strip().lower()
    query = (request.form.get("query") or "").strip()
    if not query:
        return nutrition_redirect(client.id, err="Enter a food or barcode to search.", logged_for=logged_for)

    try:
        if source == "usda":
            results = fetch_usda_foods(query)
        elif source == "openfoodfacts":
            results = fetch_open_food_facts_foods(query)
        else:
            return nutrition_redirect(client.id, err="Unknown food source.", logged_for=logged_for)
    except RuntimeError as exc:
        return nutrition_redirect(client.id, err=str(exc), logged_for=logged_for)
    except (HTTPError, URLError, TimeoutError, json.JSONDecodeError):
        return nutrition_redirect(client.id, err="Could not reach the external food source right now.", logged_for=logged_for)

    set_nutrition_search_results(client.id, results)
    if not results:
        return nutrition_redirect(client.id, msg="No matching foods found.", logged_for=logged_for)
    return nutrition_redirect(client.id, msg=f"Loaded {len(results)} result(s).", logged_for=logged_for)


@app.route("/client/<int:client_id>/nutrition/import", methods=["POST"])
@login_required
def import_food_search_result(client_id):
    if not is_admin() and current_client_id() != client_id:
        return "Forbidden", 403

    client = get_or_404(Client, client_id)
    logged_for = parse_iso_date(request.form.get("nutrition_date")) or date.today()
    try:
        idx = int(request.form.get("result_index") or "-1")
    except ValueError:
        idx = -1
    results = get_nutrition_search_results(client.id)
    if idx < 0 or idx >= len(results):
        return nutrition_redirect(client.id, err="That search result is no longer available.", logged_for=logged_for)

    item = results[idx]
    existing = (
        FoodItem.query
        .filter(FoodItem.client_id.is_(None))
        .filter_by(source=item["source"], source_ref=item.get("source_ref") or None)
        .first()
    )
    if existing:
        return nutrition_redirect(client.id, msg="That food is already in the library.", logged_for=logged_for)

    food = FoodItem(
        client_id=None,
        name=item["name"][:120],
        brand=(item.get("brand") or None),
        source=item["source"],
        source_ref=(item.get("source_ref") or None),
        barcode=(item.get("barcode") or None),
        calories_per_100g=item["calories_per_100g"],
        protein_per_100g=item["protein_per_100g"],
        carbs_per_100g=item["carbs_per_100g"],
        fat_per_100g=item["fat_per_100g"],
    )
    db.session.add(food)
    db.session.commit()
    return nutrition_redirect(client.id, msg="Food imported into the shared library.", logged_for=logged_for)


@app.route("/client/<int:client_id>/nutrition/logs/add", methods=["POST"])
@login_required
def add_food_log_entry(client_id):
    if not is_admin() and current_client_id() != client_id:
        return "Forbidden", 403

    client = get_or_404(Client, client_id)
    logged_for = parse_iso_date(request.form.get("logged_for")) or date.today()

    food_id_raw = (request.form.get("food_id") or "").strip()
    if not food_id_raw:
        search_name = normalize_food_search_text(request.form.get("food_name") or "")
        if search_name:
            matched_food = next(
                (
                    item for item in (
                        FoodItem.query
                        .filter(or_(FoodItem.client_id.is_(None), FoodItem.client_id == client.id))
                        .order_by(FoodItem.name.asc(), FoodItem.brand.asc(), FoodItem.id.asc())
                        .all()
                    )
                    if search_name in normalize_food_search_text(f"{item.name} {item.brand or ''}")
                ),
                None,
            )
            food_id_raw = str(matched_food.id) if matched_food else ""
    try:
        food_id = int(food_id_raw or "0")
    except ValueError:
        food_id = 0
    food = get_accessible_food(client.id, food_id)
    if not food:
        return nutrition_redirect(client.id, err="Select a food from your library first.", logged_for=logged_for)

    quantity_grams, err = parse_positive_float(request.form.get("quantity_grams"), "Quantity in grams")
    if err:
        return nutrition_redirect(client.id, err=err, logged_for=logged_for)

    factor = quantity_grams / 100
    food_label = food.name if not food.brand else f"{food.name} ({food.brand})"
    log_entry = FoodLogEntry(
        client_id=client.id,
        food_id=food.id,
        logged_for=logged_for,
        meal_type=normalize_meal_type(request.form.get("meal_type")),
        quantity_grams=quantity_grams,
        food_name=food_label[:160],
        calories=round(food.calories_per_100g * factor, 1),
        protein=round(food.protein_per_100g * factor, 1),
        carbs=round(food.carbs_per_100g * factor, 1),
        fat=round(food.fat_per_100g * factor, 1),
        note=((request.form.get("note") or "").strip() or None),
    )
    db.session.add(log_entry)
    db.session.commit()
    return nutrition_redirect(client.id, msg="Food logged.", logged_for=logged_for)


@app.route("/client/<int:client_id>/nutrition/log-food-ajax", methods=["POST"])
@login_required
def add_food_log_entry_ajax(client_id):
    if not is_admin() and current_client_id() != client_id:
        return jsonify({"success": False, "error": "Forbidden"}), 403

    client = get_or_404(Client, client_id)
    logged_for = parse_iso_date(request.form.get("logged_for")) or date.today()

    food_id_raw = (request.form.get("food_id") or "").strip()
    if not food_id_raw:
        search_name = normalize_food_search_text(request.form.get("food_name") or "")
        if search_name:
            matched_food = next(
                (
                    item for item in (
                        FoodItem.query
                        .filter(or_(FoodItem.client_id.is_(None), FoodItem.client_id == client.id))
                        .order_by(FoodItem.name.asc(), FoodItem.brand.asc(), FoodItem.id.asc())
                        .all()
                    )
                    if search_name in normalize_food_search_text(f"{item.name} {item.brand or ''}")
                ),
                None,
            )
            food_id_raw = str(matched_food.id) if matched_food else ""
    try:
        food_id = int(food_id_raw or "0")
    except ValueError:
        food_id = 0
    food = get_accessible_food(client.id, food_id)
    if not food:
        return jsonify({"success": False, "error": "Select a food from your library first."}), 400

    quantity_grams, err = parse_positive_float(request.form.get("quantity_grams"), "Quantity in grams")
    if err:
        return jsonify({"success": False, "error": err}), 400

    factor = quantity_grams / 100
    food_label = food.name if not food.brand else f"{food.name} ({food.brand})"
    log_entry = FoodLogEntry(
        client_id=client.id,
        food_id=food.id,
        logged_for=logged_for,
        meal_type=normalize_meal_type(request.form.get("meal_type")),
        quantity_grams=quantity_grams,
        food_name=food_label[:160],
        calories=round(food.calories_per_100g * factor, 1),
        protein=round(food.protein_per_100g * factor, 1),
        carbs=round(food.carbs_per_100g * factor, 1),
        fat=round(food.fat_per_100g * factor, 1),
        note=((request.form.get("note") or "").strip() or None),
    )
    db.session.add(log_entry)
    db.session.commit()

    set_nutrition_label_draft(client.id, None)

    # Get updated nutrition summary
    food_logs = (
        FoodLogEntry.query.filter_by(client_id=client.id, logged_for=logged_for)
        .order_by(FoodLogEntry.created_at.desc(), FoodLogEntry.id.desc())
        .all()
    )
    nutrition_summary = build_nutrition_summary(food_logs, client.daily_calorie_target)

    return jsonify({
        "success": True,
        "message": "Food logged successfully!",
        "nutrition_summary": nutrition_summary_payload(nutrition_summary),
    })


@app.route("/client/<int:client_id>/nutrition/scan-label", methods=["POST"])
@login_required
def scan_food_label(client_id):
    if not is_admin() and current_client_id() != client_id:
        return "Forbidden", 403

    client = get_or_404(Client, client_id)
    logged_for = parse_iso_date(request.form.get("nutrition_date")) or date.today()
    if not label_scan_enabled():
        return nutrition_redirect(
            client.id,
            err="Label OCR is not enabled yet. Add GOOGLE_VISION_API_KEY or install Tesseract OCR first.",
            logged_for=logged_for,
        )

    label_file = request.files.get("label_photo")
    if not label_file or not label_file.filename:
        return nutrition_redirect(client.id, err="Choose a nutrition label image first.", logged_for=logged_for)
    if not allowed_photo_file(label_file.filename):
        return nutrition_redirect(client.id, err="Use a JPG, PNG, or WEBP image for label scanning.", logged_for=logged_for)

    label_name = (request.form.get("label_name") or "").strip()
    brand = (request.form.get("label_brand") or "").strip() or None
    safe_name = secure_filename(label_file.filename)
    unique_name = f"{uuid.uuid4().hex}_{safe_name}"
    target_path = os.path.join(app.config["UPLOAD_LABEL_DIR"], unique_name)
    label_file.save(target_path)

    try:
        text = scan_label_file_text(target_path)
    except RuntimeError as exc:
        return nutrition_redirect(client.id, err=str(exc), logged_for=logged_for)
    except Exception:
        return nutrition_redirect(client.id, err="Could not read that label image.", logged_for=logged_for)
    finally:
        try:
            os.remove(target_path)
        except OSError:
            pass

    parsed = parse_label_text(text)
    if not label_name:
        label_name = os.path.splitext(safe_name)[0].replace("_", " ").strip() or "Custom Food"

    if not is_probable_nutrition_label(text, parsed):
        draft = build_manual_food_draft(label_name, brand, text)
        set_nutrition_label_draft(client.id, draft)
        return nutrition_redirect(
            client.id,
            err="No nutrition label found. Enter the macros manually below.",
            logged_for=logged_for,
            anchor="customFoodForm",
        )

    draft = build_nutrition_label_draft(label_name, brand, parsed, text)
    set_nutrition_label_draft(client.id, draft)

    if parsed["calories"] is None and parsed["protein"] is None and parsed["carbs"] is None and parsed["fat"] is None:
        return nutrition_redirect(client.id, err="Could not read the label cleanly. Review the manual form below and fill in the missing values.", logged_for=logged_for)
    return nutrition_redirect(client.id, msg="Label scanned. Review the custom food form below and save the values.", logged_for=logged_for, anchor="customFoodForm")


@app.route("/client/<int:client_id>/nutrition/scan-label-ajax", methods=["POST"])
@login_required
def scan_food_label_ajax(client_id):
    if not is_admin() and current_client_id() != client_id:
        return jsonify({"success": False, "error": "Forbidden"}), 403

    client = get_or_404(Client, client_id)
    if not label_scan_enabled():
        return jsonify({
            "success": False,
            "error": "Label OCR is not enabled yet. Add GOOGLE_VISION_API_KEY or install Tesseract OCR first.",
        }), 400

    label_file = request.files.get("label_photo")
    if not label_file or not label_file.filename:
        return jsonify({"success": False, "error": "Choose a nutrition label image first."}), 400
    if not allowed_photo_file(label_file.filename):
        return jsonify({"success": False, "error": "Use a JPG, PNG, or WEBP image for label scanning."}), 400

    label_name = (request.form.get("label_name") or "").strip()
    brand = (request.form.get("label_brand") or "").strip() or None
    safe_name = secure_filename(label_file.filename)
    unique_name = f"{uuid.uuid4().hex}_{safe_name}"
    target_path = os.path.join(app.config["UPLOAD_LABEL_DIR"], unique_name)
    label_file.save(target_path)

    try:
        text = scan_label_file_text(target_path)
    except RuntimeError as exc:
        return jsonify({"success": False, "error": str(exc)}), 400
    except Exception:
        return jsonify({"success": False, "error": "Could not read that label image."}), 400
    finally:
        try:
            os.remove(target_path)
        except OSError:
            pass

    parsed = parse_label_text(text)
    if not label_name:
        label_name = os.path.splitext(safe_name)[0].replace("_", " ").strip() or "Custom Food"

    if not is_probable_nutrition_label(text, parsed):
        draft = build_manual_food_draft(label_name, brand, text)
        set_nutrition_label_draft(client.id, draft)
        return jsonify({
            "success": True,
            "message": "No nutrition label found. Enter the macros manually.",
            "draft": draft,
            "label_found": False,
            "needs_review": True,
        })

    draft = build_nutrition_label_draft(label_name, brand, parsed, text)
    set_nutrition_label_draft(client.id, draft)
    has_any_value = any(parsed[key] is not None for key in ("calories", "protein", "carbs", "fat"))
    return jsonify({
        "success": True,
        "message": "Label scanned. Review and save the food.",
        "draft": draft,
        "label_found": True,
        "needs_review": not has_any_value,
    })


@app.route("/client/<int:client_id>/nutrition/logs/delete/<int:log_id>", methods=["POST"])
@login_required
def delete_food_log_entry(client_id, log_id):
    if not is_admin() and current_client_id() != client_id:
        return "Forbidden", 403

    log_entry = get_or_404(FoodLogEntry, log_id)
    if log_entry.client_id != client_id:
        abort(404)
    logged_for = log_entry.logged_for
    db.session.delete(log_entry)
    db.session.commit()
    return nutrition_redirect(client_id, msg="Logged food removed.", logged_for=logged_for)


@app.route("/client/<int:client_id>/strength/import", methods=["POST"])
@login_required
def import_strength_data(client_id):
    if not is_admin():
        return "Forbidden", 403

    client = get_or_404(Client, client_id)
    csv_file = request.files.get("strength_csv")
    if not csv_file or not csv_file.filename:
        return redirect(url_for("client_profile", client_id=client.id, tab="strength", err="Choose a strength export file first."))

    try:
        result = import_strength_export(client.id, csv_file, secure_filename(csv_file.filename))
    except Exception:
        return redirect(url_for("client_profile", client_id=client.id, tab="strength", err="Could not import that strength export. Use a Lyfta CSV or Hevy Excel/CSV export."))

    source_label = {"lyfta": "Lyfta", "hevy": "Hevy"}.get(result["source"], "strength")
    return redirect(url_for(
        "client_profile",
        client_id=client.id,
        tab="strength",
        msg=f"Imported {result['created']} strength sets from {source_label}." + (f" Skipped {result['skipped']} rows." if result["skipped"] else ""),
    ))


@app.route("/client/<int:client_id>/strength/poster.png")
@login_required
def strength_poster_png(client_id):
    if not is_admin():
        return "Forbidden", 403

    client = get_or_404(Client, client_id)
    strength_entries = (
        StrengthLogEntry.query.filter_by(client_id=client.id)
        .order_by(StrengthLogEntry.logged_at.asc(), StrengthLogEntry.id.asc())
        .all()
    )
    start_date, end_date = parse_strength_date_range(None, None, strength_entries)
    manual_lifts = request.args.getlist("poster_lifts")
    weight_points = (
        Measurement.query.filter_by(client_id=client.id)
        .filter(Measurement.weight.isnot(None))
        .order_by(Measurement.date.asc())
        .all()
    )
    poster = build_strength_poster_data(strength_entries, weight_points, start_date, end_date, manual_lifts)
    image_data = render_strength_poster_png(client.name, poster)
    filename = f"{secure_filename(client.name or 'client')}_strength_poster.png"
    return send_file(image_data, mimetype="image/png", as_attachment=True, download_name=filename)


@app.route("/client/<int:client_id>/change-password", methods=["POST"], endpoint="change_client_password")
@login_required
def change_client_password(client_id):
    # self-service for client accounts only
    if is_admin():
        return "Forbidden", 403
    if current_client_id() != client_id:
        return "Forbidden", 403

    user = User.query.filter_by(
        id=session.get("user_id"),
        role="client",
        client_id=client_id
    ).first()
    if not user:
        return redirect(url_for("client_profile", client_id=client_id, tab="info", err="Client user not found."))

    current_password = request.form.get("current_password") or ""
    new_username = (request.form.get("username") or "").strip()
    new_password = request.form.get("new_password") or ""
    confirm_password = request.form.get("confirm_password") or ""

    if not new_username:
        return redirect(url_for("client_profile", client_id=client_id, tab="info", err="Username is required."))
    if len(new_username) > 80:
        return redirect(url_for("client_profile", client_id=client_id, tab="info", err="Username is too long."))
    if not re.fullmatch(r"[A-Za-z0-9._-]+", new_username):
        return redirect(url_for("client_profile", client_id=client_id, tab="info", err="Username can only contain letters, numbers, dot, underscore, and hyphen."))

    changing_username = new_username != user.username
    changing_password = bool(new_password or confirm_password)
    if changing_username or changing_password or user.must_change_password:
        if not current_password:
            return redirect(url_for("client_profile", client_id=client_id, tab="info", err="Current password is required to update account settings."))
        if not check_password_hash(user.password_hash, current_password):
            return redirect(url_for("client_profile", client_id=client_id, tab="info", err="Current password is incorrect."))

    if changing_username:
        existing_user = User.query.filter_by(username=new_username).first()
        if existing_user and existing_user.id != user.id:
            return redirect(url_for("client_profile", client_id=client_id, tab="info", err="Username is already taken."))
        user.username = new_username

    if changing_password or user.must_change_password:
        if len(new_password) < 6:
            return redirect(url_for("client_profile", client_id=client_id, tab="info", err="New password must be at least 6 characters."))
        if new_password != confirm_password:
            return redirect(url_for("client_profile", client_id=client_id, tab="info", err="New password and confirmation do not match."))
        user.password_hash = generate_password_hash(new_password)
        user.must_change_password = False

    db.session.commit()

    return redirect(url_for("client_profile", client_id=client_id, tab="info", msg="Account settings updated."))



# =========================
# Stats
# =========================
@app.route("/client/<int:client_id>/stats/add", methods=["POST"], endpoint="add_measurement")
@login_required
def add_measurement(client_id):
    # client can add stats to self, admin can add to any
    if not is_admin() and current_client_id() != client_id:
        return "Forbidden", 403

    client = get_or_404(Client, client_id)

    parsed_values, err = parse_measurement_form(request.form)
    if err:
        return redirect(url_for("client_profile", client_id=client.id, tab="stats", err=err))

    m = Measurement(client_id=client.id, **parsed_values)

    db.session.add(m)
    db.session.commit()
    updated_goals = 0
    if parsed_values.get("weight") is not None:
        updated_goals = sync_weight_goal_progress(client.id)
        if updated_goals:
            db.session.commit()

    msg = "Measurement saved."
    if updated_goals:
        msg += f" Goal progress updated ({updated_goals})."
    return redirect(url_for("client_profile", client_id=client.id, tab="stats", msg=msg))


@app.route("/client/<int:client_id>/stats/update/<int:measurement_id>", methods=["POST"], endpoint="update_measurement")
@login_required
def update_measurement(client_id, measurement_id):
    if not is_admin() and current_client_id() != client_id:
        return "Forbidden", 403

    client = get_or_404(Client, client_id)
    m = get_or_404(Measurement, measurement_id)
    if m.client_id != client.id:
        abort(404)

    parsed_values, err = parse_measurement_form(request.form)
    if err:
        return redirect(url_for("client_profile", client_id=client.id, tab="stats", err=err))

    for field, value in parsed_values.items():
        setattr(m, field, value)

    db.session.commit()
    updated_goals = 0
    if parsed_values.get("weight") is not None:
        updated_goals = sync_weight_goal_progress(client.id)
        if updated_goals:
            db.session.commit()

    msg = "Measurement updated."
    if updated_goals:
        msg += f" Goal progress updated ({updated_goals})."
    return redirect(url_for("client_profile", client_id=client.id, tab="stats", msg=msg))


@app.route("/client/<int:client_id>/stats/delete/<int:measurement_id>", methods=["POST"], endpoint="delete_measurement")
@login_required
def delete_measurement(client_id, measurement_id):
    if not is_admin():
        return "Forbidden", 403

    m = get_or_404(Measurement, measurement_id)
    if m.client_id != client_id:
        abort(404)
    db.session.delete(m)
    db.session.commit()
    return redirect(url_for("client_profile", client_id=client_id, tab="stats"))


@app.route("/client/<int:client_id>/photos/upload", methods=["POST"])
@login_required
def upload_progress_photo(client_id):
    if not is_admin() and current_client_id() != client_id:
        return "Forbidden", 403
    client = get_or_404(Client, client_id)
    f = request.files.get("photo")
    image_error = validate_uploaded_image(f)
    if image_error:
        return redirect(url_for("client_profile", client_id=client.id, tab="stats", err=image_error))
    ext = os.path.splitext(f.filename)[1].lower()
    safe_name = secure_filename(f.filename)
    unique_name = f"{client.id}_{uuid.uuid4().hex}_{safe_name}"
    if ext and not unique_name.lower().endswith(ext):
        unique_name += ext
    target_path = os.path.join(app.config["UPLOAD_PROGRESS_DIR"], unique_name)
    f.save(target_path)
    note = (request.form.get("note") or "").strip()
    p = ProgressPhoto(
        client_id=client.id,
        file_name=unique_name,
        note=note,
        uploaded_by_role="admin" if is_admin() else "client",
    )
    db.session.add(p)
    db.session.commit()
    return redirect(url_for("client_profile", client_id=client.id, tab="stats", msg="Photo uploaded."))


@app.route("/client/<int:client_id>/photos/delete/<int:photo_id>", methods=["POST"])
@login_required
def delete_progress_photo(client_id, photo_id):
    if not is_admin() and current_client_id() != client_id:
        return "Forbidden", 403
    photo = get_or_404(ProgressPhoto, photo_id)
    if photo.client_id != client_id:
        abort(404)
    if not is_admin() and photo.uploaded_by_role != "client":
        return "Forbidden", 403
    path = os.path.join(app.config["UPLOAD_PROGRESS_DIR"], photo.file_name)
    if os.path.exists(path):
        try:
            os.remove(path)
        except Exception:
            pass
    db.session.delete(photo)
    db.session.commit()
    return redirect(url_for("client_profile", client_id=client_id, tab="stats", msg="Photo deleted."))


# =========================
# Sessions
# =========================
@app.route("/client/<int:client_id>/appointments/add", methods=["POST"])
@login_required
def add_appointment(client_id):
    if not is_admin() and current_client_id() != client_id:
        return "Forbidden", 403
    client = get_or_404(Client, client_id)
    dt = parse_datetime_local(request.form.get("scheduled_for"))
    if not dt:
        return redirect(url_for("client_profile", client_id=client.id, tab="info", err="Invalid appointment date/time."))
    note = (request.form.get("note") or "").strip()
    a = Appointment(
        client_id=client.id,
        scheduled_for=dt,
        status="requested",
        note=note,
        created_by_role="admin" if is_admin() else "client",
    )
    db.session.add(a)
    db.session.commit()
    return redirect(url_for("client_profile", client_id=client.id, tab="info", msg="Appointment saved."))


@app.route("/client/<int:client_id>/appointments/respond/<int:appointment_id>", methods=["POST"])
@login_required
def respond_appointment(client_id, appointment_id):
    if is_admin() or current_client_id() != client_id:
        return "Forbidden", 403

    a = get_or_404(Appointment, appointment_id)
    if a.client_id != client_id:
        abort(404)

    status = (request.form.get("status") or "").strip().lower()
    if status not in ("confirmed", "cancelled"):
        return redirect(url_for("client_profile", client_id=client_id, tab="info", err="Invalid appointment response."))
    if a.status != "requested":
        return redirect(url_for("client_profile", client_id=client_id, tab="info", err="Appointment is not awaiting response."))

    cancel_reason = (request.form.get("cancel_reason") or "").strip()
    if status == "cancelled":
        if not cancel_reason:
            return redirect(url_for("client_profile", client_id=client_id, tab="info", err="Please provide a cancellation reason."))
        reason_note = f"Client cancellation reason: {cancel_reason}"
        if a.note:
            a.note = f"{a.note} | {reason_note}"
        else:
            a.note = reason_note

    a.status = status
    db.session.commit()
    return redirect(url_for("client_profile", client_id=client_id, tab="info", msg=f"Appointment {status}."))


@app.route("/client/<int:client_id>/appointments/status/<int:appointment_id>", methods=["POST"])
@login_required
def update_appointment_status(client_id, appointment_id):
    if not is_admin():
        return "Forbidden", 403
    a = get_or_404(Appointment, appointment_id)
    if a.client_id != client_id:
        abort(404)
    status = (request.form.get("status") or "").strip().lower()
    if status not in ("requested", "confirmed", "completed", "cancelled"):
        return redirect(url_for("client_profile", client_id=client_id, tab="info", err="Invalid appointment status."))
    a.status = status
    db.session.commit()
    return redirect(url_for("client_profile", client_id=client_id, tab="info", msg="Appointment status updated."))


@app.route("/client/<int:client_id>/appointments/delete/<int:appointment_id>", methods=["POST"])
@login_required
def delete_appointment(client_id, appointment_id):
    if not is_admin():
        return "Forbidden", 403
    a = get_or_404(Appointment, appointment_id)
    if a.client_id != client_id:
        abort(404)
    db.session.delete(a)
    db.session.commit()
    return redirect(url_for("client_profile", client_id=client_id, tab="info", msg="Appointment deleted."))


@app.route("/client/<int:client_id>/sessions/add", methods=["POST"], endpoint="add_session")
@login_required
def add_session(client_id):
    # client can add sessions to self, admin can add to any
    if not is_admin() and current_client_id() != client_id:
        return "Forbidden", 403

    client = get_or_404(Client, client_id)

    sessions_per_week_from_payments, _status = get_current_plan(client.id)
    sessions_per_week = (
        sessions_per_week_from_payments
        if sessions_per_week_from_payments is not None
        else (client.weekly_sessions or 0)
    )

    session_date_raw = (request.form.get("session_date") or "").strip()
    session_day = parse_iso_date(session_date_raw) if session_date_raw else date.today()
    if not session_day:
        return redirect(url_for(
            "client_profile",
            client_id=client.id,
            tab="info",
            err="Invalid session date."
        ))
    if session_day > date.today():
        return redirect(url_for(
            "client_profile",
            client_id=client.id,
            tab="info",
            err="Session date cannot be in the future."
        ))

    target_week_start = week_start(session_day)
    used_this_week = (
        SessionLog.query.filter_by(client_id=client.id)
        .filter(SessionLog.date >= datetime(target_week_start.year, target_week_start.month, target_week_start.day))
        .filter(SessionLog.date < datetime(target_week_start.year, target_week_start.month, target_week_start.day) + timedelta(days=7))
        .count()
    )
    bonus = 0
    if client.rollover_for_week == target_week_start and (client.rollover_bonus or 0) > 0:
        bonus = client.rollover_bonus
    allowed = max((sessions_per_week or 0) + bonus, 0)
    remaining = max(allowed - used_this_week, 0)

    if remaining <= 0:
        return redirect(url_for(
            "client_profile",
            client_id=client.id,
            tab="info",
            err=f"Weekly limit reached for week of {target_week_start.strftime('%d/%m/%Y')} ({used_this_week}/{allowed})."
        ))

    note = (request.form.get("note") or "").strip()
    session_dt = datetime(session_day.year, session_day.month, session_day.day, 12, 0, 0)
    s = SessionLog(client_id=client.id, note=note, date=session_dt)
    db.session.add(s)
    db.session.commit()

    return redirect(
        url_for(
            "client_profile",
            client_id=client.id,
            tab="info",
            msg=f"Session logged for {session_day.strftime('%d/%m/%Y')}"
        )
    )


@app.route("/client/<int:client_id>/sessions/delete/<int:session_id>", methods=["POST"], endpoint="delete_session")
@login_required
def delete_session(client_id, session_id):
    if not is_admin():
        return "Forbidden", 403

    s = get_or_404(SessionLog, session_id)
    if s.client_id != client_id:
        abort(404)
    db.session.delete(s)
    db.session.commit()
    return redirect(url_for("client_profile", client_id=client_id, tab="info"))


@app.route("/client/<int:client_id>/sessions/transfer", methods=["POST"], endpoint="transfer_sessions")
@login_required
def transfer_sessions(client_id):
    # trainer action only (admin)
    if not is_admin():
        return "Forbidden", 403

    client = get_or_404(Client, client_id)

    sessions_per_week_from_payments, _status = get_current_plan(client.id)
    sessions_per_week = (
        sessions_per_week_from_payments
        if sessions_per_week_from_payments is not None
        else (client.weekly_sessions or 0)
    )

    today = date.today()
    ws = week_start(today)

    if client.last_transfer_week == ws:
        return redirect(url_for(
            "client_profile", client_id=client.id, tab="info",
            err="Transfer already done for this week."
        ))

    used_this_week, remaining, bonus, allowed = compute_sessions(client, sessions_per_week)

    # remaining is what can be transferred
    transfer_amount = remaining

    # apply bonus to NEXT week
    next_week = ws + timedelta(days=7)
    client.rollover_bonus = transfer_amount
    client.rollover_for_week = next_week
    client.last_transfer_week = ws

    db.session.commit()

    return redirect(url_for(
        "client_profile", client_id=client.id, tab="info",
        msg=f"Transferred {transfer_amount} unused session(s) to next week."
    ))


# =========================
# Payments
# =========================
@app.route("/client/<int:client_id>/payments/add", methods=["POST"], endpoint="add_payment")
@login_required
def add_payment(client_id):
    if not is_admin():
        return "Forbidden", 403

    client = get_or_404(Client, client_id)

    start = parse_ddmmyyyy(request.form.get("start_date"))
    if not start:
        return redirect(url_for("client_profile", client_id=client.id, tab="payments", err="Invalid start date. Use DD/MM/YYYY."))

    amount_paid = to_int(request.form.get("amount_paid"), default=0)
    note = (request.form.get("note") or "").strip()

    plan_type = (request.form.get("plan_type") or "").strip()
    if amount_paid <= 0:
        return redirect(url_for("client_profile", client_id=client.id, tab="payments", err="Enter a valid amount."))

    if plan_type == "custom":
        months = to_int(request.form.get("custom_months"), default=0)
        sessions_per_week = to_int(request.form.get("custom_sessions_per_week"), default=0)
        if months <= 0:
            return redirect(url_for(
                "client_profile", client_id=client.id, tab="payments",
                err="Enter a valid length (in months) for the custom plan."
            ))
        if sessions_per_week <= 0:
            return redirect(url_for(
                "client_profile", client_id=client.id, tab="payments",
                err="Enter a valid number of weekly sessions for the custom plan."
            ))
        monthly_price = amount_paid // months
    elif plan_type in ("5000", "7000"):
        monthly_price = int(plan_type)
        sessions_per_week = 3 if monthly_price == 5000 else 5
        if amount_paid % monthly_price != 0:
            return redirect(url_for(
                "client_profile", client_id=client.id, tab="payments",
                err=f"Amount must be a multiple of {monthly_price} for the selected plan."
            ))
        months = amount_paid // monthly_price
    else:
        # Backwards-compatible fallback for older forms/API calls.
        if amount_paid % 5000 == 0 and amount_paid % 7000 != 0:
            monthly_price = 5000
            sessions_per_week = 3
            months = amount_paid // 5000
        elif amount_paid % 7000 == 0 and amount_paid % 5000 != 0:
            monthly_price = 7000
            sessions_per_week = 5
            months = amount_paid // 7000
        elif amount_paid % 5000 == 0 and amount_paid % 7000 == 0:
            return redirect(url_for(
                "client_profile", client_id=client.id, tab="payments",
                err="Choose a plan type for this amount."
            ))
        else:
            return redirect(url_for(
                "client_profile", client_id=client.id, tab="payments",
                err=f"Amount must be divisible by 5000 or 7000. (Received: {amount_paid})"
            ))

    p = Payment(
        client_id=client.id,
        start_date=start,
        months=months,
        monthly_price=monthly_price,
        sessions_per_week=sessions_per_week,
        amount_paid=amount_paid,
        note=note
    )

    db.session.add(p)
    db.session.commit()

    return redirect(url_for("client_profile", client_id=client.id, tab="payments", msg="Payment saved"))


@app.route("/client/<int:client_id>/payments/adjust-current", methods=["POST"], endpoint="adjust_current_payment_due")
@login_required
def adjust_current_payment_due(client_id):
    if not is_admin():
        return "Forbidden", 403

    client = get_or_404(Client, client_id)
    latest_payment = (
        Payment.query.filter_by(client_id=client.id)
        .order_by(Payment.start_date.desc(), Payment.paid_on.desc(), Payment.id.desc())
        .first()
    )
    if not latest_payment:
        return redirect(url_for("client_profile", client_id=client.id, tab="payments", err="No payment plan to adjust."))

    due_date_raw = (request.form.get("new_due_date") or "").strip()
    extend_days_raw = (request.form.get("extend_days") or "").strip()
    if not due_date_raw and not extend_days_raw:
        return redirect(url_for("client_profile", client_id=client.id, tab="payments", err="Enter days to extend or a new due date."))

    new_due_date = None
    if due_date_raw:
        new_due_date = parse_ddmmyyyy(due_date_raw)
        if not new_due_date:
            return redirect(url_for("client_profile", client_id=client.id, tab="payments", err="Invalid due date. Use DD/MM/YYYY."))

    if extend_days_raw:
        extend_days = to_int(extend_days_raw, default=0)
        if extend_days <= 0:
            return redirect(url_for("client_profile", client_id=client.id, tab="payments", err="Extend days must be greater than zero."))
        if new_due_date:
            return redirect(url_for("client_profile", client_id=client.id, tab="payments", err="Use either extend days or a due date, not both."))
        base_due = payment_due_date(latest_payment)
        new_due_date = base_due + timedelta(days=extend_days)

    latest_payment.due_date_override = new_due_date
    db.session.commit()
    return redirect(url_for("client_profile", client_id=client.id, tab="payments", msg="Current plan due date updated."))


@app.route("/client/<int:client_id>/payments/delete/<int:payment_id>", methods=["POST"], endpoint="delete_payment")
@login_required
def delete_payment(client_id, payment_id):
    if not is_admin():
        return "Forbidden", 403

    p = get_or_404(Payment, payment_id)
    if p.client_id != client_id:
        abort(404)
    db.session.delete(p)
    db.session.commit()
    return redirect(url_for("client_profile", client_id=client_id, tab="payments"))


# =========================
# Admin: Create client login
# =========================
@app.route("/client/<int:client_id>/create-login", methods=["POST"], endpoint="create_client_login")
@login_required
def create_client_login(client_id):
    if not is_admin():
        return "Forbidden", 403

    client = get_or_404(Client, client_id)
    username = (request.form.get("username") or "").strip()
    password = request.form.get("password") or ""

    if not username or not password:
        return redirect(url_for("client_profile", client_id=client.id, tab="info", err="Username and password required."))

    if User.query.filter_by(username=username).first():
        return redirect(url_for("client_profile", client_id=client.id, tab="info", err="Username already taken."))
    if User.query.filter_by(client_id=client.id).filter(User.role != "admin").first():
        return redirect(url_for("client_profile", client_id=client.id, tab="info", err="Client login already exists. Use reset/deactivate tools."))

    u = User(
        username=username,
        password_hash=generate_password_hash(password),
        role="client",
        client_id=client.id,
        must_change_password=True,
    )
    db.session.add(u)
    db.session.commit()

    return redirect(url_for("client_profile", client_id=client.id, tab="info", msg="Client login created."))


@app.route("/client/<int:client_id>/login/reset-password", methods=["POST"], endpoint="admin_reset_client_password")
@login_required
def admin_reset_client_password(client_id):
    if not is_admin():
        return "Forbidden", 403

    client = get_or_404(Client, client_id)
    user = (
        User.query
        .filter_by(client_id=client.id)
        .filter(User.role != "admin")
        .order_by(User.id.desc())
        .first()
    )
    if not user:
        return redirect(url_for("client_profile", client_id=client.id, tab="info", err="Client login does not exist yet."))

    new_password = request.form.get("new_password") or ""
    if len(new_password) < 6:
        return redirect(url_for("client_profile", client_id=client.id, tab="info", err="Temporary password must be at least 6 characters."))

    user.password_hash = generate_password_hash(new_password)
    user.must_change_password = True
    db.session.commit()
    return redirect(url_for("client_profile", client_id=client.id, tab="info", msg="Client password reset."))


@app.route("/client/<int:client_id>/login/deactivate", methods=["POST"], endpoint="deactivate_client_login")
@login_required
def deactivate_client_login(client_id):
    if not is_admin():
        return "Forbidden", 403

    client = get_or_404(Client, client_id)
    user = (
        User.query
        .filter_by(client_id=client.id)
        .filter(User.role != "admin")
        .order_by(User.id.desc())
        .first()
    )
    if not user:
        return redirect(url_for("client_profile", client_id=client.id, tab="info", err="Client login does not exist yet."))
    if user.role == "disabled":
        return redirect(url_for("client_profile", client_id=client.id, tab="info", msg="Client login is already deactivated."))

    user.role = "disabled"
    db.session.commit()
    return redirect(url_for("client_profile", client_id=client.id, tab="info", msg="Client login deactivated."))


@app.route("/client/<int:client_id>/login/reactivate", methods=["POST"], endpoint="reactivate_client_login")
@login_required
def reactivate_client_login(client_id):
    if not is_admin():
        return "Forbidden", 403

    client = get_or_404(Client, client_id)
    user = (
        User.query
        .filter_by(client_id=client.id)
        .filter(User.role != "admin")
        .order_by(User.id.desc())
        .first()
    )
    if not user:
        return redirect(url_for("client_profile", client_id=client.id, tab="info", err="Client login does not exist yet."))
    if user.role != "disabled":
        return redirect(url_for("client_profile", client_id=client.id, tab="info", msg="Client login is already active."))

    user.role = "client"
    db.session.commit()
    return redirect(url_for("client_profile", client_id=client.id, tab="info", msg="Client login reactivated."))


# =========================
# Debug
# =========================
@app.route("/routes")
@login_required
def show_routes():
    if not is_admin():
        return "Forbidden", 403
    return "<br>".join(sorted([f"{r.endpoint} -> {r.rule}" for r in app.url_map.iter_rules()]))


@app.route("/ping")
def ping():
    return "PING OK"


# =========================
# Start
# =========================
if __name__ == "__main__":
    with app.app_context():
        db.create_all()
        seed_admin()

    debug_mode = truthy(os.environ.get("FLASK_DEBUG", "0")) and not IS_PROD
    host = os.environ.get("HOST", "127.0.0.1")
    port = int(os.environ.get("PORT", "5000"))
    app.run(host=host, port=port, debug=debug_mode, use_reloader=debug_mode)
